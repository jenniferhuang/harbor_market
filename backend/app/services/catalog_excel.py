from __future__ import annotations

import hashlib
import json
import logging
import re
import zipfile
from collections import defaultdict
from collections.abc import Callable
from dataclasses import dataclass
from datetime import UTC, datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Any
from uuid import uuid4

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from pydantic import ValidationError
from sqlalchemy import select
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session, selectinload

from app.core.errors import ApiError
from app.models import (
    Category,
    ImportJob,
    ObjectCleanupJob,
    Product,
    ProductImage,
    ProductSku,
    User,
)
from app.schemas.catalog import ImportErrorItem, ProductCreate, ProductSkuInput
from app.services.object_cleanup import IMPORT_LEASE, enqueue_object_cleanup
from app.services.object_storage import (
    ObjectStorage,
    ObjectStorageNotFoundError,
    ObjectStorageSizeError,
)
from app.services.product_images import ProductImageValidationError, inspect_product_image
from app.services.text_validation import validate_xml_safe_text

PRODUCT_HEADERS = [
    "product_code",
    "name",
    "subtitle",
    "category_code",
    "status",
    "base_price_yuan",
    "market_price_yuan",
    "unit",
    "stock_status",
    "inventory_count",
    "featured",
    "sort_order",
    "tags",
    "selling_points",
    "description",
    "ingredients",
    "allergen_info",
    "specifications_json",
]
SKU_HEADERS = [
    "product_code",
    "sku_code",
    "name",
    "price_yuan",
    "market_price_yuan",
    "stock_quantity",
    "attributes_json",
    "is_default",
    "is_active",
    "sort_order",
]
IMAGE_HEADERS = ["product_code", "image_type", "object_key", "alt_text", "sort_order"]

_PRODUCT_VALIDATION_FIELD_ALIASES = {
    "base_price_cents": "base_price_yuan",
    "market_price_cents": "market_price_yuan",
    "category_id": "category_code",
    "specifications": "specifications_json",
    "skus": "is_default",
}
_SKU_VALIDATION_FIELD_ALIASES = {
    "price_cents": "price_yuan",
    "market_price_cents": "market_price_yuan",
    "attributes": "attributes_json",
}

_TEXT_HEADERS_BY_SHEET = {
    "Products": frozenset({"product_code", "category_code"}),
    "SKUs": frozenset({"product_code", "sku_code"}),
    "Images": frozenset({"product_code", "object_key"}),
}
_PRODUCT_CODE_KEY_PATTERN = r"[A-Z0-9][A-Z0-9_.-]{0,63}"
_IMAGE_FILE_KEY_PATTERN = r"[0-9a-f]{32}\.(?:jpg|png|webp)"
_STAGED_IMAGE_KEY_RE = re.compile(
    rf"products/staged/(?P<product_code>{_PRODUCT_CODE_KEY_PATTERN})/"
    rf"(?P<filename>{_IMAGE_FILE_KEY_PATTERN})"
)
_CATALOG_IMAGE_KEY_RE = re.compile(
    rf"products/catalog/(?P<product_code>{_PRODUCT_CODE_KEY_PATTERN})/"
    rf"(?P<image_type>cover|gallery|detail)/(?P<filename>{_IMAGE_FILE_KEY_PATTERN})"
)
_DIRECT_IMAGE_KEY_RE = re.compile(
    rf"products/(?P<product_id>[1-9][0-9]*)/(?P<image_type>cover|gallery|detail)/"
    rf"(?P<filename>{_IMAGE_FILE_KEY_PATTERN})"
)

_HEADER_FILL = PatternFill("solid", fgColor="173F37")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_SUBTLE_FILL = PatternFill("solid", fgColor="EDF3F0")
_ERROR_FILL = PatternFill("solid", fgColor="FCE8E6")
_MAX_IMPORT_ERRORS = 500
_MAX_IMPORT_IMAGE_ROWS = 500
_MAX_IMPORT_IMAGE_BYTES = 100 * 1024 * 1024
_MAX_XLSX_ENTRIES = 2_000
_MAX_XLSX_UNCOMPRESSED_BYTES = 64 * 1024 * 1024
_MAX_XLSX_ENTRY_BYTES = 32 * 1024 * 1024
_MAX_XLSX_COMPRESSION_RATIO = 1_000
_PG_INTEGER_MIN = -(2**31)
_PG_INTEGER_MAX = 2**31 - 1
logger = logging.getLogger(__name__)


class CatalogImportStorageError(RuntimeError):
    """Signals a retryable object-storage failure during catalog import."""


class WorkbookFieldError(ValueError):
    def __init__(self, field: str, message: str) -> None:
        super().__init__(message)
        self.field = field


class SkuWorkbookFieldError(WorkbookFieldError):
    def __init__(self, sku_code: str, field: str, message: str) -> None:
        super().__init__(field, message)
        self.sku_code = sku_code


@dataclass(slots=True)
class ParsedImage:
    row: int
    product_code: str
    image_type: str
    object_key: str
    alt_text: str | None
    sort_order: int
    mime_type: str | None
    size_bytes: int
    width: int
    height: int
    extension: str
    etag: str | None
    is_staged: bool


@dataclass(slots=True)
class ParsedWorkbook:
    products: list[ProductCreate]
    images: list[ParsedImage]


@dataclass(frozen=True, slots=True)
class PromotedImage:
    source_key: str
    destination_key: str
    cleanup_job_id: int


def create_template(categories: list[Category]) -> bytes:
    workbook = Workbook()
    products = workbook.active
    products.title = "Products"
    skus = workbook.create_sheet("SKUs")
    images = workbook.create_sheet("Images")
    dictionary = workbook.create_sheet("Dictionary")
    _prepare_data_sheet(products, PRODUCT_HEADERS)
    _prepare_data_sheet(skus, SKU_HEADERS)
    _prepare_data_sheet(images, IMAGE_HEADERS)
    _add_validations(products, skus, images)
    _populate_dictionary(dictionary, categories)
    return _save_workbook(workbook)


def export_catalog(session: Session) -> bytes:
    categories = list(session.scalars(select(Category).order_by(Category.sort_order, Category.id)))
    products = list(
        session.scalars(
            select(Product)
            .options(
                selectinload(Product.category),
                selectinload(Product.skus),
                selectinload(Product.images),
            )
            .order_by(Product.sort_order, Product.id)
        )
    )
    workbook = Workbook()
    products_sheet = workbook.active
    products_sheet.title = "Products"
    skus_sheet = workbook.create_sheet("SKUs")
    images_sheet = workbook.create_sheet("Images")
    dictionary = workbook.create_sheet("Dictionary")
    _prepare_data_sheet(products_sheet, PRODUCT_HEADERS)
    _prepare_data_sheet(skus_sheet, SKU_HEADERS)
    _prepare_data_sheet(images_sheet, IMAGE_HEADERS)

    for product in products:
        _append_safe_row(
            products_sheet,
            [
                _excel_safe(product.product_code),
                _excel_safe(product.name),
                _excel_safe(product.subtitle),
                product.category.code,
                product.status,
                Decimal(product.base_price_cents) / 100,
                (
                    Decimal(product.market_price_cents) / 100
                    if product.market_price_cents is not None
                    else None
                ),
                _excel_safe(product.unit),
                product.stock_status,
                product.inventory_count,
                product.featured,
                product.sort_order,
                _excel_safe(json.dumps(product.tags, ensure_ascii=False)),
                _excel_safe(json.dumps(product.selling_points, ensure_ascii=False)),
                _excel_safe(product.description),
                _excel_safe(product.ingredients),
                _excel_safe(product.allergen_info),
                _excel_safe(json.dumps(product.specifications, ensure_ascii=False)),
            ],
        )
        for sku in sorted(product.skus, key=lambda item: (item.sort_order, item.id)):
            _append_safe_row(
                skus_sheet,
                [
                    product.product_code,
                    sku.sku_code,
                    _excel_safe(sku.name),
                    Decimal(sku.price_cents) / 100,
                    (
                        Decimal(sku.market_price_cents) / 100
                        if sku.market_price_cents is not None
                        else None
                    ),
                    sku.stock_quantity,
                    _excel_safe(json.dumps(sku.attributes, ensure_ascii=False)),
                    sku.is_default,
                    sku.is_active,
                    sku.sort_order,
                ],
            )
        for image in sorted(product.images, key=lambda item: (item.image_type, item.sort_order)):
            _append_safe_row(
                images_sheet,
                [
                    product.product_code,
                    image.image_type,
                    _excel_safe(image.object_key),
                    _excel_safe(image.alt_text),
                    image.sort_order,
                ],
            )
    for sheet, price_columns in ((products_sheet, (6, 7)), (skus_sheet, (4, 5))):
        for column in price_columns:
            for row in range(2, sheet.max_row + 1):
                sheet.cell(row, column).number_format = "¥0.00"
    _add_validations(products_sheet, skus_sheet, images_sheet)
    _populate_dictionary(dictionary, categories)
    return _save_workbook(workbook)


def import_catalog(
    session: Session,
    *,
    user: User,
    filename: str,
    payload: bytes,
    dry_run: bool,
    storage: ObjectStorage,
    image_max_bytes: int = 5 * 1024 * 1024,
    idempotency_key: str | None = None,
) -> ImportJob:
    workbook_sha256 = hashlib.sha256(payload).hexdigest()
    if idempotency_key is not None:
        existing_job = session.scalar(
            select(ImportJob).where(
                ImportJob.created_by == user.id,
                ImportJob.idempotency_key == idempotency_key,
            )
        )
        if existing_job is not None:
            return _resolve_idempotent_import(
                existing_job,
                workbook_sha256=workbook_sha256,
                dry_run=dry_run,
            )
    job = ImportJob(
        creator=user,
        original_filename=(filename or "products.xlsx")[:255],
        workbook_sha256=workbook_sha256,
        idempotency_key=idempotency_key,
        dry_run=dry_run,
        status="pending",
    )
    session.add(job)
    try:
        session.commit()
    except IntegrityError:
        session.rollback()
        if idempotency_key is None:
            raise
        existing_job = session.scalar(
            select(ImportJob).where(
                ImportJob.created_by == user.id,
                ImportJob.idempotency_key == idempotency_key,
            )
        )
        if existing_job is None:
            raise
        return _resolve_idempotent_import(
            existing_job,
            workbook_sha256=workbook_sha256,
            dry_run=dry_run,
        )
    job_id = job.id

    promoted_images: list[PromotedImage] = []
    staging_cleanup_jobs: list[ObjectCleanupJob] = []
    try:
        parsed, errors, summary = parse_workbook(
            session,
            payload,
            storage,
            image_max_bytes=image_max_bytes,
        )
    except CatalogImportStorageError:
        logger.warning("Object storage unavailable while parsing import job %s", job_id)
        return _fail_import_job(
            session,
            job_id,
            summary=_summary(0, 0, 0, 1),
            message="对象存储暂不可用，请稍后重试",
        )
    except Exception:
        logger.exception("Unexpected failure while parsing catalog import job %s", job_id)
        return _fail_import_job(
            session,
            job_id,
            summary=_summary(0, 0, 0, 1),
            message="Excel 解析失败，导入任务已安全终止",
        )
    if errors:
        return _fail_import_job(session, job_id, summary=summary, errors=errors)
    if dry_run:
        job.status = "validated"
        job.summary = summary
        job.errors = []
        job.completed_at = datetime.now(UTC)
        session.commit()
        return job

    promotion_plans: list[PromotedImage] = []
    try:
        promotion_plans = _prepare_promotion_intents(
            session,
            parsed.images,
            created_by=user.id,
        )
        affected_codes = {item.product_code for item in parsed.products} | {
            item.product_code for item in parsed.images
        }
        _lock_affected_products(session, affected_codes)
        # The first pass gives fast row-level feedback. Reparse while the affected
        # product locks are held so image/SKU validation observes the final state.
        parsed, errors, summary = parse_workbook(
            session,
            payload,
            storage,
            image_max_bytes=image_max_bytes,
            validated_images={item.object_key: item for item in parsed.images},
        )
        if errors:
            return _fail_import_job(
                session,
                job_id,
                summary=summary,
                errors=errors,
                cleanup_job_ids=[item.cleanup_job_id for item in promotion_plans],
            )
        _apply_import(
            session,
            parsed,
            storage,
            promoted_images,
            staging_cleanup_jobs,
            promotion_plans,
        )
        job = session.get(ImportJob, job_id)
        if job is None:
            raise RuntimeError("Import job disappeared during transaction")
        job.status = "completed"
        job.promoted_staging_keys = [item.source_key for item in promoted_images]
        job.summary = {
            **summary,
            "cleanup_queued": len(staging_cleanup_jobs),
        }
        job.errors = []
        job.completed_at = datetime.now(UTC)
        session.commit()
    except CatalogImportStorageError:
        return _fail_import_job(
            session,
            job_id,
            summary=summary,
            message="对象存储暂不可用，导入事务已回滚，请稍后重试",
            cleanup_job_ids=[item.cleanup_job_id for item in promotion_plans],
        )
    except (IntegrityError, ValueError):
        return _fail_import_job(
            session,
            job_id,
            summary=summary,
            message="导入数据与现有记录冲突，整个事务已回滚",
            cleanup_job_ids=[item.cleanup_job_id for item in promotion_plans],
        )
    except Exception:
        logger.exception("Unexpected failure while applying catalog import job %s", job_id)
        return _fail_import_job(
            session,
            job_id,
            summary=summary,
            message="导入执行失败，整个事务已回滚",
            cleanup_job_ids=[item.cleanup_job_id for item in promotion_plans],
        )
    return job


def _resolve_idempotent_import(
    job: ImportJob,
    *,
    workbook_sha256: str,
    dry_run: bool,
) -> ImportJob:
    if job.workbook_sha256 != workbook_sha256 or job.dry_run != dry_run:
        raise ApiError(
            409,
            "idempotency_key_conflict",
            "The idempotency key was already used for a different import",
        )
    if job.status == "pending":
        raise ApiError(
            409,
            "import_in_progress",
            "The original import request is still in progress; retry with the same key later",
        )
    return job


def _prepare_promotion_intents(
    session: Session,
    images: list[ParsedImage],
    *,
    created_by: int,
) -> list[PromotedImage]:
    staged_images = [item for item in images if item.is_staged]
    destinations = [
        (f"products/catalog/{item.product_code}/{item.image_type}/{uuid4().hex}.{item.extension}")
        for item in staged_images
    ]
    cleanup_jobs = enqueue_object_cleanup(
        session,
        destinations,
        reason="import_promotion_intent",
        status="intent",
        # A bounded import can involve hundreds of copies. Keep live imports out
        # of the stale-intent worker while still recovering after process death.
        not_before=datetime.now(UTC) + IMPORT_LEASE,
        created_by=created_by,
    )
    session.commit()
    return [
        PromotedImage(
            source_key=image.object_key,
            destination_key=destination,
            cleanup_job_id=cleanup_job.id,
        )
        for image, destination, cleanup_job in zip(
            staged_images,
            destinations,
            cleanup_jobs,
            strict=True,
        )
    ]


def _fail_import_job(
    session: Session,
    job_id: int,
    *,
    summary: dict[str, int],
    errors: list[ImportErrorItem] | None = None,
    message: str | None = None,
    cleanup_job_ids: list[int] | None = None,
) -> ImportJob:
    session.rollback()
    job = session.get(ImportJob, job_id)
    if job is None:
        raise RuntimeError("Import job disappeared while recording failure")
    job.status = "failed"
    recorded_errors = [
        error.model_dump()
        for error in (
            errors
            or [
                ImportErrorItem(
                    sheet="Workbook",
                    row=0,
                    field="transaction",
                    message=message or "导入失败",
                )
            ]
        )
    ]
    cleanup_jobs = list(
        session.scalars(
            select(ObjectCleanupJob).where(ObjectCleanupJob.id.in_(cleanup_job_ids or [-1]))
        )
    )
    for cleanup_job in cleanup_jobs:
        if cleanup_job.status != "completed":
            cleanup_job.status = "pending"
            cleanup_job.not_before = None
    active_cleanup_ids = [job.id for job in cleanup_jobs if job.status != "completed"]
    if active_cleanup_ids:
        recorded_errors.append(
            ImportErrorItem(
                sheet="ObjectStorage",
                row=0,
                field="cleanup",
                message=(
                    "导入已回滚；复制对象清理已排队，任务 ID："
                    + ",".join(str(item) for item in active_cleanup_ids[:20])
                ),
            ).model_dump()
        )
    job.summary = {
        **summary,
        **({"cleanup_queued": len(active_cleanup_ids)} if active_cleanup_ids else {}),
    }
    job.errors = recorded_errors
    job.completed_at = datetime.now(UTC)
    session.commit()
    return job


def _lock_affected_products(session: Session, product_codes: set[str]) -> None:
    if not product_codes:
        return
    list(
        session.scalars(
            select(Product.id)
            .where(Product.product_code.in_(product_codes))
            .order_by(Product.id)
            .with_for_update()
        )
    )
    # Keep the database locks, but discard identity-map relationships loaded by
    # the first validation pass before re-reading the authoritative state.
    session.expire_all()


def parse_workbook(
    session: Session,
    payload: bytes,
    storage: ObjectStorage,
    *,
    image_max_bytes: int = 5 * 1024 * 1024,
    validated_images: dict[str, ParsedImage] | None = None,
) -> tuple[ParsedWorkbook, list[ImportErrorItem], dict[str, int]]:
    errors: list[ImportErrorItem] = []
    if not payload:
        errors.append(_error("Workbook", 0, "file", "Excel 文件为空"))
        return ParsedWorkbook([], []), errors, _summary(0, 0, 0, len(errors))
    try:
        _validate_xlsx_archive(payload)
    except ValueError as exc:
        errors.append(_error("Workbook", 0, "file", str(exc)))
        return ParsedWorkbook([], []), errors, _summary(0, 0, 0, len(errors))
    try:
        workbook = load_workbook(BytesIO(payload), read_only=True, data_only=True)
    except Exception:
        errors.append(_error("Workbook", 0, "file", "文件不是有效的 .xlsx 工作簿"))
        return ParsedWorkbook([], []), errors, _summary(0, 0, 0, len(errors))

    required_sheets = {"Products", "SKUs", "Images"}
    for sheet_name in sorted(required_sheets - set(workbook.sheetnames)):
        errors.append(_error(sheet_name, 1, "sheet", "缺少必需的工作表"))
    if errors:
        workbook.close()
        return ParsedWorkbook([], []), errors, _summary(0, 0, 0, len(errors))

    product_rows = _read_rows(workbook["Products"], PRODUCT_HEADERS, errors, limit=5_000)
    sku_rows = _read_rows(workbook["SKUs"], SKU_HEADERS, errors, limit=20_000)
    image_rows = _read_rows(
        workbook["Images"],
        IMAGE_HEADERS,
        errors,
        limit=_MAX_IMPORT_IMAGE_ROWS,
    )
    workbook.close()

    categories = {item.code: item for item in session.scalars(select(Category))}
    existing_products = {
        item.product_code: item
        for item in session.scalars(
            select(Product).options(selectinload(Product.images), selectinload(Product.skus))
        )
    }
    existing_skus = {item.sku_code: item.product_id for item in session.scalars(select(ProductSku))}
    existing_images = {
        item.object_key: item.product_id for item in session.scalars(select(ProductImage))
    }

    sku_inputs, sku_source_rows = _parse_skus(sku_rows, errors)
    declared_product_codes = {
        _text(row.get("product_code")).upper() for _row_number, row in product_rows
    }
    known_product_codes = declared_product_codes | set(existing_products)
    for row_number, row in sku_rows:
        product_code = _text(row.get("product_code")).upper()
        if product_code and product_code not in known_product_codes:
            errors.append(
                _error(
                    "SKUs",
                    row_number,
                    "product_code",
                    "SKU 引用的商品编码不存在",
                )
            )
    products = _parse_products(
        product_rows,
        sku_inputs,
        sku_source_rows,
        categories,
        existing_products,
        errors,
    )
    product_codes = {item.product_code for item in products}
    known_codes = product_codes | set(existing_products)
    _validate_sku_ownership(
        products,
        existing_skus,
        existing_products,
        sku_source_rows,
        errors,
    )
    images = _parse_images(
        image_rows,
        known_codes,
        existing_products,
        existing_images,
        storage,
        errors,
        image_max_bytes=image_max_bytes,
        validated_images=validated_images or {},
    )
    _validate_image_limits(products, images, existing_products, errors)
    error_count = len(errors)
    errors = _capped_errors(errors)
    summary = _summary(
        len(products), sum(len(item.skus) for item in products), len(images), error_count
    )
    return ParsedWorkbook(products, images), errors, summary


def _parse_products(
    rows: list[tuple[int, dict[str, Any]]],
    sku_inputs: dict[str, list[ProductSkuInput]],
    sku_source_rows: dict[str, int],
    categories: dict[str, Category],
    existing_products: dict[str, Product],
    errors: list[ImportErrorItem],
) -> list[ProductCreate]:
    products: list[ProductCreate] = []
    seen: set[str] = set()
    for row_number, row in rows:
        code = _text(row.get("product_code")).upper()
        if not code:
            errors.append(_error("Products", row_number, "product_code", "此字段必填"))
            continue
        if code in seen:
            errors.append(_error("Products", row_number, "product_code", "商品编码重复"))
            continue
        seen.add(code)
        category_code = _text(row.get("category_code")).upper()
        category = categories.get(category_code)
        if category is None:
            errors.append(_error("Products", row_number, "category_code", "类目编码不存在"))
            continue
        try:
            base_price = _parse_excel_field(
                "base_price_yuan",
                _yuan_to_cents,
                row.get("base_price_yuan"),
                required=True,
            )
            market_price = _parse_excel_field(
                "market_price_yuan",
                _yuan_to_cents,
                row.get("market_price_yuan"),
                required=False,
            )
            inventory_count = _parse_excel_field(
                "inventory_count",
                _integer,
                row.get("inventory_count"),
                default=None,
            )
            product_skus = _final_skus(
                existing_products.get(code),
                sku_inputs.get(code, []),
                default_code=f"{code}-DEFAULT",
                default_price_cents=base_price,
                default_market_price_cents=market_price,
                default_stock_quantity=inventory_count or 0,
            )
            product = ProductCreate(
                product_code=code,
                name=_text(row.get("name")),
                subtitle=_optional_text(row.get("subtitle")),
                category_id=category.id,
                status=_text(row.get("status")) or "draft",
                base_price_cents=base_price,
                market_price_cents=market_price,
                unit=_text(row.get("unit")) or "件",
                stock_status=_text(row.get("stock_status")) or "in_stock",
                inventory_count=inventory_count,
                featured=_parse_excel_field(
                    "featured",
                    _boolean,
                    row.get("featured"),
                    default=False,
                ),
                sort_order=(
                    _parse_excel_field(
                        "sort_order",
                        _integer,
                        row.get("sort_order"),
                        default=0,
                    )
                    or 0
                ),
                tags=_parse_excel_field("tags", _text_list, row.get("tags")),
                selling_points=_parse_excel_field(
                    "selling_points",
                    _text_list,
                    row.get("selling_points"),
                ),
                description=_text(row.get("description")),
                ingredients=_optional_text(row.get("ingredients")),
                allergen_info=_optional_text(row.get("allergen_info")),
                specifications=_parse_excel_field(
                    "specifications_json",
                    _json_value,
                    row.get("specifications_json"),
                    expected=list,
                ),
                skus=product_skus,
            )
        except SkuWorkbookFieldError as exc:
            errors.append(
                _error(
                    "SKUs",
                    sku_source_rows.get(exc.sku_code, 0),
                    exc.field,
                    str(exc),
                )
            )
            continue
        except WorkbookFieldError as exc:
            errors.append(_error("Products", row_number, exc.field, str(exc)))
            continue
        except ValidationError as exc:
            errors.append(
                _error(
                    "Products",
                    row_number,
                    _validation_error_field(exc, _PRODUCT_VALIDATION_FIELD_ALIASES),
                    _validation_message(exc),
                )
            )
            continue
        if product.status == "published" and not category.is_active:
            errors.append(
                _error(
                    "Products",
                    row_number,
                    "category_code",
                    "已上架商品必须属于启用类目",
                )
            )
            continue
        products.append(product)

    for code, incoming_skus in sku_inputs.items():
        if code in seen:
            continue
        existing = existing_products.get(code)
        if existing is None:
            continue
        try:
            merged_skus = _final_skus(
                existing,
                incoming_skus,
                default_code=f"{code}-DEFAULT",
                default_price_cents=existing.base_price_cents,
                default_market_price_cents=existing.market_price_cents,
                default_stock_quantity=existing.inventory_count or 0,
            )
            products.append(
                ProductCreate(
                    product_code=existing.product_code,
                    name=existing.name,
                    subtitle=existing.subtitle,
                    category_id=existing.category_id,
                    status=existing.status,
                    base_price_cents=existing.base_price_cents,
                    market_price_cents=existing.market_price_cents,
                    currency=existing.currency,
                    unit=existing.unit,
                    description=existing.description,
                    featured=existing.featured,
                    stock_status=existing.stock_status,
                    inventory_count=existing.inventory_count,
                    tags=list(existing.tags),
                    selling_points=list(existing.selling_points),
                    specifications=list(existing.specifications),
                    ingredients=existing.ingredients,
                    allergen_info=existing.allergen_info,
                    sort_order=existing.sort_order,
                    skus=merged_skus,
                )
            )
        except SkuWorkbookFieldError as exc:
            errors.append(
                _error(
                    "SKUs",
                    sku_source_rows.get(exc.sku_code, 0),
                    exc.field,
                    str(exc),
                )
            )
        except WorkbookFieldError as exc:
            source_row = next(
                (sku_source_rows.get(sku.sku_code, 0) for sku in incoming_skus),
                0,
            )
            errors.append(_error("SKUs", source_row, exc.field, str(exc)))
        except ValidationError as exc:
            source_row = next(
                (sku_source_rows.get(sku.sku_code, 0) for sku in incoming_skus),
                0,
            )
            errors.append(
                _error(
                    "SKUs",
                    source_row,
                    _validation_error_field(exc, _SKU_VALIDATION_FIELD_ALIASES),
                    _validation_message(exc),
                )
            )
    return products


def _parse_skus(
    rows: list[tuple[int, dict[str, Any]]],
    errors: list[ImportErrorItem],
) -> tuple[dict[str, list[ProductSkuInput]], dict[str, int]]:
    by_product: dict[str, list[ProductSkuInput]] = defaultdict(list)
    source_rows: dict[str, int] = {}
    seen: set[str] = set()
    for row_number, row in rows:
        product_code = _text(row.get("product_code")).upper()
        sku_code = _text(row.get("sku_code")).upper()
        if not product_code or not sku_code:
            errors.append(
                _error("SKUs", row_number, "product_code/sku_code", "商品编码和 SKU 编码均为必填")
            )
            continue
        if sku_code in seen:
            errors.append(_error("SKUs", row_number, "sku_code", "SKU 编码重复"))
            continue
        seen.add(sku_code)
        try:
            sku = ProductSkuInput(
                sku_code=sku_code,
                name=_text(row.get("name")),
                price_cents=_parse_excel_field(
                    "price_yuan",
                    _yuan_to_cents,
                    row.get("price_yuan"),
                    required=True,
                ),
                market_price_cents=_parse_excel_field(
                    "market_price_yuan",
                    _yuan_to_cents,
                    row.get("market_price_yuan"),
                    required=False,
                ),
                stock_quantity=(
                    _parse_excel_field(
                        "stock_quantity",
                        _integer,
                        row.get("stock_quantity"),
                        default=0,
                    )
                    or 0
                ),
                attributes=_parse_excel_field(
                    "attributes_json",
                    _json_value,
                    row.get("attributes_json"),
                    expected=dict,
                ),
                is_default=_parse_excel_field(
                    "is_default",
                    _boolean,
                    row.get("is_default"),
                    default=False,
                ),
                is_active=_parse_excel_field(
                    "is_active",
                    _boolean,
                    row.get("is_active"),
                    default=True,
                ),
                sort_order=(
                    _parse_excel_field(
                        "sort_order",
                        _integer,
                        row.get("sort_order"),
                        default=0,
                    )
                    or 0
                ),
            )
        except WorkbookFieldError as exc:
            errors.append(_error("SKUs", row_number, exc.field, str(exc)))
            continue
        except ValidationError as exc:
            errors.append(
                _error(
                    "SKUs",
                    row_number,
                    _validation_error_field(exc, _SKU_VALIDATION_FIELD_ALIASES),
                    _validation_message(exc),
                )
            )
            continue
        by_product[product_code].append(sku)
        source_rows[sku.sku_code] = row_number
    return by_product, source_rows


def _final_skus(
    existing_product: Product | None,
    incoming_skus: list[ProductSkuInput],
    *,
    default_code: str,
    default_price_cents: int,
    default_market_price_cents: int | None,
    default_stock_quantity: int,
) -> list[ProductSkuInput]:
    merged: dict[str, ProductSkuInput] = {}
    if existing_product is not None:
        merged = {
            sku.sku_code: ProductSkuInput(
                sku_code=sku.sku_code,
                name=sku.name,
                price_cents=sku.price_cents,
                market_price_cents=sku.market_price_cents,
                stock_quantity=sku.stock_quantity,
                attributes=dict(sku.attributes),
                is_default=sku.is_default,
                is_active=sku.is_active,
                sort_order=sku.sort_order,
            )
            for sku in existing_product.skus
        }

    incoming_defaults = [sku for sku in incoming_skus if sku.is_active and sku.is_default]
    if len(incoming_defaults) > 1:
        raise SkuWorkbookFieldError(
            incoming_defaults[1].sku_code,
            "is_default",
            "每个商品必须且只能有一个启用的默认 SKU",
        )
    if incoming_defaults:
        for sku in merged.values():
            sku.is_default = False
    for incoming in incoming_skus:
        merged[incoming.sku_code] = incoming.model_copy(deep=True)

    if not merged:
        merged[default_code] = ProductSkuInput(
            sku_code=default_code,
            name="默认规格",
            price_cents=default_price_cents,
            market_price_cents=default_market_price_cents,
            stock_quantity=default_stock_quantity,
            is_default=True,
        )

    active = [sku for sku in merged.values() if sku.is_active]
    if not active:
        source = incoming_skus[-1] if incoming_skus else next(iter(merged.values()))
        raise SkuWorkbookFieldError(
            source.sku_code,
            "is_active",
            "商品至少需要一个启用的 SKU",
        )
    defaults = [sku for sku in active if sku.is_default]
    if not defaults:
        active[0].is_default = True
    elif len(defaults) > 1:
        incoming_codes = {sku.sku_code for sku in incoming_skus}
        source = next(
            (sku for sku in reversed(defaults) if sku.sku_code in incoming_codes),
            defaults[1],
        )
        raise SkuWorkbookFieldError(
            source.sku_code,
            "is_default",
            "每个商品必须且只能有一个启用的默认 SKU",
        )
    return sorted(merged.values(), key=lambda sku: (sku.sort_order, sku.sku_code))


def _parse_images(
    rows: list[tuple[int, dict[str, Any]]],
    known_codes: set[str],
    existing_products: dict[str, Product],
    existing_images: dict[str, int],
    storage: ObjectStorage,
    errors: list[ImportErrorItem],
    *,
    image_max_bytes: int,
    validated_images: dict[str, ParsedImage],
) -> list[ParsedImage]:
    images: list[ParsedImage] = []
    seen: set[str] = set()
    total_image_bytes = 0
    for row_number, row in rows:
        product_code = _text(row.get("product_code")).upper()
        object_key = _text(row.get("object_key"))
        image_type = _text(row.get("image_type")).lower()
        if product_code not in known_codes:
            errors.append(_error("Images", row_number, "product_code", "商品编码不存在"))
            continue
        if not object_key:
            errors.append(_error("Images", row_number, "object_key", "此字段必填"))
            continue
        if object_key in seen:
            errors.append(_error("Images", row_number, "object_key", "对象路径重复"))
            continue
        seen.add(object_key)
        if not _is_safe_object_key(object_key):
            errors.append(
                _error("Images", row_number, "object_key", "对象路径格式不安全或长度超限")
            )
            continue
        if image_type not in {"cover", "gallery", "detail"}:
            errors.append(
                _error("Images", row_number, "image_type", "只能填写 cover、gallery 或 detail")
            )
            continue
        owner_id = existing_images.get(object_key)
        expected_owner = existing_products.get(product_code)
        if owner_id is not None and (expected_owner is None or owner_id != expected_owner.id):
            errors.append(_error("Images", row_number, "object_key", "该对象已关联到其他商品"))
            continue
        key_kind = _image_object_key_kind(
            object_key,
            product_code=product_code,
            allow_catalog_product_mismatch=owner_id is not None,
        )
        if key_kind is None:
            errors.append(
                _error(
                    "Images",
                    row_number,
                    "object_key",
                    "对象路径必须是与商品编码和图片用途匹配的 staging 或正式商品图片路径",
                )
            )
            continue
        try:
            stat = storage.stat(object_key)
        except ObjectStorageNotFoundError:
            errors.append(_error("Images", row_number, "object_key", "对象存储中不存在该路径"))
            continue
        except Exception as exc:
            raise CatalogImportStorageError("object storage stat failed") from exc
        if stat.size < 1 or stat.size > image_max_bytes:
            errors.append(
                _error(
                    "Images",
                    row_number,
                    "object_key",
                    f"图片大小必须在 1 到 {image_max_bytes} 字节之间",
                )
            )
            continue
        total_image_bytes += stat.size
        if total_image_bytes > _MAX_IMPORT_IMAGE_BYTES:
            errors.append(
                _error(
                    "Images",
                    row_number,
                    "object_key",
                    f"单次导入图片总大小不能超过 {_MAX_IMPORT_IMAGE_BYTES} 字节",
                )
            )
            break
        cached = validated_images.get(object_key)
        if (
            cached is not None
            and cached.size_bytes == stat.size
            and cached.etag is not None
            and cached.etag == stat.etag
        ):
            mime_type = cached.mime_type
            extension = cached.extension
            width = cached.width
            height = cached.height
        else:
            try:
                payload = storage.get(object_key, max_bytes=image_max_bytes)
            except ObjectStorageSizeError:
                errors.append(_error("Images", row_number, "object_key", "图片实际大小超过上限"))
                continue
            except ObjectStorageNotFoundError:
                errors.append(_error("Images", row_number, "object_key", "对象存储中不存在该路径"))
                continue
            except Exception as exc:
                raise CatalogImportStorageError("object storage read failed") from exc
            if len(payload) != stat.size:
                errors.append(
                    _error("Images", row_number, "object_key", "图片实际大小与对象元数据不一致")
                )
                continue
            try:
                inspection = inspect_product_image(payload)
            except ProductImageValidationError:
                errors.append(
                    _error(
                        "Images",
                        row_number,
                        "object_key",
                        "对象内容不是有效的 JPEG、PNG 或 WebP 图片",
                    )
                )
                continue
            mime_type = inspection.mime_type
            extension = inspection.extension
            width = inspection.width
            height = inspection.height
        alt_text = _optional_text(row.get("alt_text"))
        if alt_text is not None:
            try:
                validate_xml_safe_text(alt_text, label="图片说明", max_length=200)
            except ValueError:
                errors.append(
                    _error(
                        "Images",
                        row_number,
                        "alt_text",
                        "图片说明不能超过 200 个字符且不能包含控制字符",
                    )
                )
                continue
        try:
            sort_order = _integer(row.get("sort_order"), default=0) or 0
        except ValueError as exc:
            errors.append(_error("Images", row_number, "sort_order", str(exc)))
            continue
        images.append(
            ParsedImage(
                row=row_number,
                product_code=product_code,
                image_type=image_type,
                object_key=object_key,
                alt_text=alt_text,
                sort_order=sort_order,
                mime_type=mime_type,
                size_bytes=stat.size,
                width=width,
                height=height,
                extension=extension,
                etag=stat.etag,
                is_staged=key_kind == "staged",
            )
        )
    return images


def _validate_sku_ownership(
    products: list[ProductCreate],
    existing_skus: dict[str, int],
    existing_products: dict[str, Product],
    source_rows: dict[str, int],
    errors: list[ImportErrorItem],
) -> None:
    for product in products:
        existing_product = existing_products.get(product.product_code)
        for sku in product.skus:
            owner_id = existing_skus.get(sku.sku_code)
            if owner_id is not None and (
                existing_product is None or owner_id != existing_product.id
            ):
                errors.append(
                    _error(
                        "SKUs",
                        source_rows.get(sku.sku_code, 0),
                        "sku_code",
                        "SKU 编码已属于其他商品",
                    )
                )


def _validate_image_limits(
    products: list[ProductCreate],
    imported: list[ParsedImage],
    existing: dict[str, Product],
    errors: list[ImportErrorItem],
) -> None:
    imported_by_product: dict[str, list[ParsedImage]] = defaultdict(list)
    for image in imported:
        imported_by_product[image.product_code].append(image)
    limits = {"cover": 1, "gallery": 8, "detail": 20}
    imported_products = {product.product_code: product for product in products}
    affected_codes = set(imported_products) | set(imported_by_product)
    for product_code in affected_codes:
        final_roles: dict[str, str] = {}
        current = existing.get(product_code)
        if current is not None:
            for image in current.images:
                final_roles[image.object_key] = image.image_type
        for image in imported_by_product[product_code]:
            final_roles[image.object_key] = image.image_type
        counts: dict[str, int] = defaultdict(int)
        for image_type in final_roles.values():
            counts[image_type] += 1
        for image_type, limit in limits.items():
            if counts[image_type] > limit:
                errors.append(
                    _error(
                        "Images",
                        next(
                            (
                                image.row
                                for image in imported_by_product[product_code]
                                if image.image_type == image_type
                            ),
                            0,
                        ),
                        product_code,
                        f"{image_type} 图片数量超过上限 {limit}",
                    )
                )
        imported_product = imported_products.get(product_code)
        final_status = imported_product.status if imported_product is not None else current.status
        if final_status == "published" and counts["cover"] != 1:
            errors.append(
                _error(
                    "Images",
                    imported_by_product[product_code][0].row
                    if imported_by_product[product_code]
                    else 0,
                    product_code,
                    "已上架商品必须且只能有一张封面图",
                )
            )


def _apply_import(
    session: Session,
    parsed: ParsedWorkbook,
    storage: ObjectStorage,
    promoted_images: list[PromotedImage],
    staging_cleanup_jobs: list[ObjectCleanupJob],
    promotion_plans: list[PromotedImage],
) -> None:
    categories = {item.id: item for item in session.scalars(select(Category))}
    affected_codes = {item.product_code for item in parsed.products} | {
        item.product_code for item in parsed.images
    }
    existing = {
        item.product_code: item
        for item in session.scalars(
            select(Product)
            .where(Product.product_code.in_(affected_codes))
            .options(selectinload(Product.skus), selectinload(Product.images))
            .execution_options(populate_existing=True)
            .with_for_update()
        )
    }
    imported_products: dict[str, Product] = {}
    for item in parsed.products:
        product = existing.get(item.product_code)
        values = item.model_dump(exclude={"skus"})
        if product is None:
            product = Product(product_code=item.product_code)
            session.add(product)
        for field, value in values.items():
            if field != "category_id":
                setattr(product, field, value)
        product.category = categories[item.category_id]
        if product.id is not None:
            for sku in product.skus:
                if sku.is_default:
                    sku.is_default = False
            session.flush()
        current_skus = {sku.sku_code: sku for sku in product.skus}
        for sku_input in item.skus:
            sku = current_skus.pop(sku_input.sku_code, None)
            if sku is None:
                sku = ProductSku(sku_code=sku_input.sku_code)
                product.skus.append(sku)
            for field, value in sku_input.model_dump().items():
                setattr(sku, field, value)
        for stale in current_skus.values():
            product.skus.remove(stale)
            session.delete(stale)
        session.flush()
        imported_products[item.product_code] = product

    existing_image_by_key = {
        image.object_key: image for product in existing.values() for image in product.images
    }
    promotion_by_source = {item.source_key: item for item in promotion_plans}
    final_roles = {item.object_key: item.image_type for item in parsed.images}
    for object_key, image in existing_image_by_key.items():
        if image.image_type == "cover" and final_roles.get(object_key, "cover") != "cover":
            image.image_type = "gallery"
    session.flush()
    for item in parsed.images:
        product = imported_products.get(item.product_code) or existing.get(item.product_code)
        if product is None:
            raise ValueError("导入过程中找不到图片关联的商品")
        image = existing_image_by_key.get(item.object_key)
        destination_key = item.object_key
        if item.is_staged:
            promoted = promotion_by_source.get(item.object_key)
            if promoted is None:
                raise ValueError("暂存图片缺少提升计划")
            destination_key = promoted.destination_key
            promoted_images.append(promoted)
            try:
                copied = storage.copy(
                    item.object_key,
                    destination_key,
                    source_etag=item.etag,
                )
            except Exception as exc:
                raise CatalogImportStorageError("staging image promotion failed") from exc
            if copied.size != item.size_bytes:
                raise ValueError("图片提升后的大小校验失败")
        if image is None:
            image = ProductImage(product=product, object_key=destination_key)
            session.add(image)
        else:
            image.object_key = destination_key
        image.image_type = item.image_type
        image.alt_text = item.alt_text or product.name
        image.sort_order = item.sort_order
        image.mime_type = item.mime_type
        image.size_bytes = item.size_bytes
        image.width = item.width
        image.height = item.height
        if item.is_staged:
            promotion_intent = session.get(
                ObjectCleanupJob,
                promotion_by_source[item.object_key].cleanup_job_id,
            )
            if promotion_intent is None:
                raise RuntimeError("图片提升清理意图丢失")
            promotion_intent.status = "completed"
            promotion_intent.last_error = None
            promotion_intent.completed_at = datetime.now(UTC)
    promoted_source_keys = [item.source_key for item in promoted_images]
    if promoted_source_keys:
        expiry_jobs = list(
            session.scalars(
                select(ObjectCleanupJob).where(
                    ObjectCleanupJob.object_key.in_(promoted_source_keys),
                    ObjectCleanupJob.reason == "staging_expiry",
                    ObjectCleanupJob.status != "completed",
                )
            )
        )
        now = datetime.now(UTC)
        for expiry_job in expiry_jobs:
            expiry_job.status = "completed"
            expiry_job.last_error = None
            expiry_job.completed_at = now
    staging_cleanup_jobs.extend(
        enqueue_object_cleanup(
            session,
            promoted_source_keys,
            reason="staging_promoted",
        )
    )
    session.flush()


def _read_rows(
    sheet: Any,
    expected_headers: list[str],
    errors: list[ImportErrorItem],
    *,
    limit: int,
) -> list[tuple[int, dict[str, Any]]]:
    iterator = sheet.iter_rows(values_only=True)
    first = next(iterator, None)
    headers = [_text(value) for value in first] if first is not None else []
    if headers != expected_headers:
        errors.append(_error(sheet.title, 1, "headers", "工作表表头与模板不一致"))
        return []
    rows: list[tuple[int, dict[str, Any]]] = []
    for row_number, values in enumerate(iterator, start=2):
        if row_number > limit + 1:
            errors.append(_error(sheet.title, row_number, "rows", f"工作表超过 {limit} 行上限"))
            break
        if all(value is None or _text(value) == "" for value in values):
            continue
        padded = list(values) + [None] * max(0, len(headers) - len(values))
        rows.append((row_number, dict(zip(headers, padded, strict=False))))
    return rows


def _prepare_data_sheet(sheet: Any, headers: list[str]) -> None:
    sheet.append(headers)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    sheet.sheet_view.showGridLines = False
    for cell in sheet[1]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    sheet.row_dimensions[1].height = 26
    widths = {
        "product_code": 20,
        "sku_code": 22,
        "name": 24,
        "subtitle": 30,
        "description": 36,
        "specifications_json": 42,
        "attributes_json": 32,
        "object_key": 48,
        "alt_text": 26,
    }
    for index, header in enumerate(headers, start=1):
        column = sheet.column_dimensions[get_column_letter(index)]
        column.width = widths.get(header, 18)
        if header in _TEXT_HEADERS_BY_SHEET.get(sheet.title, frozenset()):
            column.number_format = "@"


def _add_validations(products: Any, skus: Any, images: Any) -> None:
    validations = [
        (products, "E2:E5001", '"draft,published,archived"'),
        (products, "I2:I5001", '"in_stock,out_of_stock,preorder"'),
        (products, "K2:K5001", '"TRUE,FALSE"'),
        (skus, "H2:I20001", '"TRUE,FALSE"'),
        (images, "B2:B501", '"cover,gallery,detail"'),
    ]
    for sheet, cell_range, formula in validations:
        validation = DataValidation(type="list", formula1=formula, allow_blank=False)
        sheet.add_data_validation(validation)
        validation.add(cell_range)
    for sheet in (products, skus, images):
        sheet.conditional_formatting.add(
            f"A2:A{max(sheet.max_row, 2)}",
            FormulaRule(formula=["LEN(A2)=0"], fill=_ERROR_FILL),
        )


def _populate_dictionary(sheet: Any, categories: list[Category]) -> None:
    sheet.sheet_view.showGridLines = False
    sheet.append(["字段/字典", "说明", "示例", "允许值"])
    rows = [
        (
            "使用流程",
            (
                "以 GET /api/v1/admin/products/template.xlsx 下载的工作簿为唯一事实来源；"
                "填写后先用 dry_run=true 校验，valid=true 后再用 dry_run=false 正式导入"
            ),
            "下载最新模板 → 填写 → dry-run → 正式导入",
            "不要复用旧模板或自行新增、改名、重排表头",
        ),
        ("金额", "Excel 使用人民币元，最多两位小数；数据库/API 使用整数分", "19.90", "非负数"),
        (
            "product_code",
            "商品稳定编码，按文本处理；字母会转为大写，保留数字编码的前导零",
            "PRODUCT-01",
            "1-64 位；字母、数字、点、下划线或连字符",
        ),
        (
            "sku_code",
            "SKU 稳定编码，按文本处理；同一工作簿和系统内不可重复",
            "PRODUCT-01-LARGE",
            "1-80 位；字母、数字、点、下划线或连字符",
        ),
        ("status", "发布状态", "draft", "draft | published | archived"),
        ("stock_status", "库存展示状态", "in_stock", "in_stock | out_of_stock | preorder"),
        (
            "image_type",
            "图片用途；上架必须且只能有 1 张 cover",
            "gallery",
            "cover | gallery | detail",
        ),
        ("tags", "推荐填写 JSON 字符串数组，最多 20 个", '["新品","热卖"]', "JSON array"),
        (
            "selling_points",
            "推荐填写 JSON 字符串数组，最多 5 个",
            '["现磨咖啡","低糖可选"]',
            "JSON array",
        ),
        (
            "specifications_json",
            "最多20组；single/multiple；每组最多50项；JSON最多20000字符",
            (
                '[{"code":"temperature","name":"温度","selection_mode":"single",'
                '"required":true,"min_select":1,"max_select":1,"options":['
                '{"code":"iced","name":"冰","price_delta_cents":0,"sort":0,'
                '"is_default":true}]}]'
            ),
            "JSON array",
        ),
        ("attributes_json", "SKU 属性 JSON 对象", '{"temperature":"iced"}', "JSON object"),
        (
            "object_key",
            "按文本处理；必须是已存在于私有 MinIO 的相对对象路径，导出路径应原样保留",
            "products/staged/PRODUCT-01/0123456789abcdef0123456789abcdef.webp",
            (
                "新图使用 products/staged/{product_code}/{32位小写十六进制}.jpg|png|webp；"
                "导出也可包含 products/{id}/{image_type}/... 或 "
                "products/catalog/{product_code}/{image_type}/...；禁止绝对路径和 .."
            ),
        ),
        (
            "图片文件",
            "工作簿不嵌入图片；先通过后台上传到私有对象存储，再把返回的 object_key 填入 Images",
            "先上传，再复制 object_key",
            "JPEG | PNG | WebP；文件内容和扩展名必须有效",
        ),
        ("图片数量", "封面 1；轮播最多 8；详情最多 20", "", "cover=1, gallery<=8, detail<=20"),
    ]
    for row in rows:
        sheet.append(row)
    sheet.append([])
    sheet.append(["category_code", "当前可用类目", "category_name", "active"])
    for category in categories:
        _append_safe_row(
            sheet,
            [category.code, category.description or "", category.name, category.is_active],
        )
    for cell in sheet[1]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
    category_header_row = len(rows) + 3
    for cell in sheet[category_header_row]:
        cell.fill = _SUBTLE_FILL
        cell.font = Font(bold=True, color="173F37")
    sheet.freeze_panes = "A2"
    sheet.column_dimensions["A"].width = 26
    sheet.column_dimensions["B"].width = 58
    sheet.column_dimensions["C"].width = 32
    sheet.column_dimensions["D"].width = 42
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def _save_workbook(workbook: Workbook) -> bytes:
    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def _append_safe_row(sheet: Any, values: list[Any]) -> None:
    """Append user-controlled strings as literal XLSX strings, never formulas."""
    sheet.append(values)
    text_headers = _TEXT_HEADERS_BY_SHEET.get(sheet.title, frozenset())
    for index, cell in enumerate(sheet[sheet.max_row], start=1):
        if isinstance(cell.value, str):
            cell.data_type = "s"
        if sheet.cell(1, index).value in text_headers:
            cell.number_format = "@"


def _validate_xlsx_archive(payload: bytes) -> None:
    try:
        with zipfile.ZipFile(BytesIO(payload)) as archive:
            entries = archive.infolist()
    except (OSError, zipfile.BadZipFile, zipfile.LargeZipFile) as exc:
        raise ValueError("文件不是有效的 .xlsx 工作簿") from exc
    if len(entries) > _MAX_XLSX_ENTRIES:
        raise ValueError("Excel 压缩包文件项数量超过安全上限")
    total_size = 0
    seen_names: set[str] = set()
    for entry in entries:
        name = entry.filename.replace("\\", "/")
        parts = name.split("/")
        if (
            not name
            or name.startswith("/")
            or any(part in {"", ".", ".."} for part in parts if not name.endswith("/"))
            or name in seen_names
            or entry.flag_bits & 0x1
        ):
            raise ValueError("Excel 压缩包包含不安全的文件项")
        seen_names.add(name)
        if entry.file_size > _MAX_XLSX_ENTRY_BYTES:
            raise ValueError("Excel 压缩包单个文件项超过安全上限")
        total_size += entry.file_size
        if total_size > _MAX_XLSX_UNCOMPRESSED_BYTES:
            raise ValueError("Excel 解压后大小超过安全上限")
        if entry.file_size and (
            entry.compress_size == 0
            or entry.file_size / entry.compress_size > _MAX_XLSX_COMPRESSION_RATIO
        ):
            raise ValueError("Excel 压缩比超过安全上限")


def _is_safe_object_key(object_key: str) -> bool:
    if (
        not object_key
        or object_key.startswith("/")
        or object_key.endswith("/")
        or "\\" in object_key
        or len(object_key) > 512
        or len(object_key.encode("utf-8")) > 1_024
    ):
        return False
    return not any(
        part in {"", ".", ".."}
        or any(ord(character) < 32 or ord(character) == 127 for character in part)
        for part in object_key.split("/")
    )


def _image_object_key_kind(
    object_key: str,
    *,
    product_code: str,
    allow_catalog_product_mismatch: bool,
) -> str | None:
    staged = _STAGED_IMAGE_KEY_RE.fullmatch(object_key)
    if staged is not None:
        return "staged" if staged.group("product_code") == product_code else None

    catalog = _CATALOG_IMAGE_KEY_RE.fullmatch(object_key)
    if catalog is not None:
        if allow_catalog_product_mismatch or catalog.group("product_code") == product_code:
            return "canonical"
        return None

    direct = _DIRECT_IMAGE_KEY_RE.fullmatch(object_key)
    if direct is not None:
        return "canonical"
    return None


def _yuan_to_cents(value: Any, *, required: bool) -> int | None:
    if value is None or _text(value) == "":
        if required:
            raise ValueError("价格为必填")
        return None
    try:
        decimal_value = Decimal(str(value))
    except (InvalidOperation, TypeError, ValueError) as exc:
        raise ValueError("价格必须是数字") from exc
    if not decimal_value.is_finite() or decimal_value < 0 or decimal_value.as_tuple().exponent < -2:
        raise ValueError("价格必须为非负数且最多两位小数")
    cents = int(decimal_value * 100)
    if cents > _PG_INTEGER_MAX:
        raise ValueError("价格超过数据库整数上限")
    return cents


def _boolean(value: Any, *, default: bool) -> bool:
    if value is None or _text(value) == "":
        return default
    if isinstance(value, bool):
        return value
    normalized = _text(value).casefold()
    if normalized in {"true", "1", "yes", "y", "是"}:
        return True
    if normalized in {"false", "0", "no", "n", "否"}:
        return False
    raise ValueError("布尔值只能填写 TRUE 或 FALSE")


def _integer(value: Any, *, default: int | None) -> int | None:
    if value is None or _text(value) == "":
        return default
    if isinstance(value, bool):
        raise ValueError("整数格式无效")
    try:
        parsed = int(value)
    except (OverflowError, TypeError, ValueError) as exc:
        raise ValueError("数值必须是整数") from exc
    if parsed < _PG_INTEGER_MIN or parsed > _PG_INTEGER_MAX:
        raise ValueError("整数超出数据库允许范围")
    if str(value).strip().endswith(".0") or parsed == value:
        return parsed
    if isinstance(value, str) and re.fullmatch(r"[+-]?\d+", value.strip()):
        return parsed
    raise ValueError("数值必须是整数")


def _json_value(value: Any, *, expected: type[list] | type[dict]) -> Any:
    if value is None or _text(value) == "":
        return expected()
    serialized = str(value)
    if len(serialized) > 30_000:
        raise ValueError("JSON 内容超过 Excel 可安全往返的长度上限")
    try:
        parsed = _strict_json_loads(serialized)
    except (RecursionError, ValueError) as exc:
        raise ValueError("JSON 格式无效") from exc
    if not isinstance(parsed, expected):
        expected_name = "数组" if expected is list else "对象"
        raise ValueError(f"JSON 必须是{expected_name}")
    return parsed


def _strict_json_loads(value: str) -> Any:
    def reject_constant(constant: str) -> None:
        raise ValueError(f"非标准 JSON 数值 {constant}")

    return json.loads(value, parse_constant=reject_constant)


def _text_list(value: Any) -> list[str]:
    if value is None:
        return []
    text = str(value).strip()
    if text.startswith("["):
        try:
            parsed = _strict_json_loads(text)
        except (RecursionError, ValueError) as exc:
            raise ValueError("文本列表 JSON 格式无效") from exc
        if not isinstance(parsed, list) or any(not isinstance(item, str) for item in parsed):
            raise ValueError("文本列表 JSON 必须是字符串数组")
        return [item.strip() for item in parsed if item.strip()]
    return [part.strip() for part in re.split(r"[|;；]", text) if part.strip()]


def _text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def _optional_text(value: Any) -> str | None:
    text = _text(value)
    return text or None


def _excel_safe(value: Any) -> Any:
    return value


def _parse_excel_field[T](
    field: str,
    operation: Callable[..., T],
    /,
    *args: Any,
    **kwargs: Any,
) -> T:
    try:
        return operation(*args, **kwargs)
    except ValidationError:
        raise
    except ValueError as exc:
        raise WorkbookFieldError(field, str(exc)) from exc


def _validation_error_field(error: ValidationError, aliases: dict[str, str]) -> str:
    first = error.errors()[0]
    root_field = next(
        (str(part) for part in first.get("loc", ()) if isinstance(part, str)),
        "row",
    )
    if root_field in {"row", "__root__"}:
        message = str(first.get("msg", ""))
        referenced_field = next(
            (
                field
                for field in sorted(aliases, key=len, reverse=True)
                if field in message
            ),
            None,
        )
        if referenced_field is not None:
            return aliases[referenced_field]
    return aliases.get(root_field, root_field)


def _validation_message(error: ValueError | ValidationError) -> str:
    if isinstance(error, ValidationError):
        first = error.errors()[0]
        field = ".".join(str(part) for part in first.get("loc", ()))
        error_type = str(first.get("type", ""))
        messages = {
            "missing": "必填字段缺失",
            "string_too_short": "文本长度过短",
            "string_too_long": "文本长度过长",
            "string_pattern_mismatch": "格式不符合要求",
            "greater_than_equal": "数值不能小于允许的最小值",
            "literal_error": "值不在允许的枚举范围内",
            "list_too_long": "列表项数量超过上限",
        }
        message = messages.get(error_type)
        if message is None:
            raw_message = str(first.get("msg", ""))
            if "market_price_cents" in raw_message:
                message = "划线价不能低于当前价格"
            elif "duplicates" in raw_message:
                message = "列表中不能包含重复值"
            else:
                message = "字段值无效"
        return f"{field}: {message}" if field else message
    return str(error)


def _error(sheet: str, row: int, field: str, message: str) -> ImportErrorItem:
    return ImportErrorItem(sheet=sheet, row=row, field=field, message=message)


def _capped_errors(errors: list[ImportErrorItem]) -> list[ImportErrorItem]:
    if len(errors) <= _MAX_IMPORT_ERRORS:
        return errors
    return [
        *errors[: _MAX_IMPORT_ERRORS - 1],
        _error(
            "Workbook",
            0,
            "errors",
            f"错误过多，仅返回前 {_MAX_IMPORT_ERRORS - 1} 条；请修正后重新校验",
        ),
    ]


def _summary(products: int, skus: int, images: int, errors: int) -> dict[str, int]:
    return {"products": products, "skus": skus, "images": images, "errors": errors}
