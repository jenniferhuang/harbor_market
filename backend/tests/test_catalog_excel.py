from __future__ import annotations

import hashlib
import re
import zipfile
from io import BytesIO
from typing import Any

import pytest
from fastapi import FastAPI
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from PIL import Image
from sqlalchemy import select

from app.models import ImportJob, Product, User
from app.services import catalog_excel

PRODUCT_HEADERS = [
    "product_code",
    "name",
    "subtitle",
    "category_code",
    "status",
    "base_price_yuan",
    "market_price_yuan",
    "unit",
    "stock_status",
    "inventory_count",
    "featured",
    "sort_order",
    "tags",
    "selling_points",
    "description",
    "ingredients",
    "allergen_info",
    "specifications_json",
]
SKU_HEADERS = [
    "product_code",
    "sku_code",
    "name",
    "price_yuan",
    "market_price_yuan",
    "stock_quantity",
    "attributes_json",
    "is_default",
    "is_active",
    "sort_order",
]
IMAGE_HEADERS = ["product_code", "image_type", "object_key", "alt_text", "sort_order"]
XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _create_category(client: TestClient, code: str = "COFFEE") -> dict[str, Any]:
    response = client.post(
        "/api/v1/admin/categories",
        json={"code": code, "name": "咖啡", "description": "现磨咖啡"},
    )
    assert response.status_code == 201, response.text
    return response.json()["data"]


def _image_bytes(image_format: str = "PNG", size: tuple[int, int] = (4, 3)) -> bytes:
    output = BytesIO()
    Image.new("RGB", size, color=(24, 86, 72)).save(output, format=image_format)
    return output.getvalue()


def _create_product_with_skus(
    client: TestClient,
    category_id: int,
    *,
    code: str,
) -> dict[str, Any]:
    response = client.post(
        "/api/v1/admin/products",
        json={
            "product_code": code,
            "name": "局部导入商品",
            "category_id": category_id,
            "base_price_cents": 1990,
            "market_price_cents": 2390,
            "skus": [
                {
                    "sku_code": f"{code}-A",
                    "name": "规格 A",
                    "price_cents": 1990,
                    "market_price_cents": 2390,
                    "stock_quantity": 10,
                    "is_default": True,
                    "is_active": True,
                    "sort_order": 1,
                },
                {
                    "sku_code": f"{code}-B",
                    "name": "规格 B",
                    "price_cents": 2090,
                    "market_price_cents": 2490,
                    "stock_quantity": 8,
                    "is_default": False,
                    "is_active": True,
                    "sort_order": 2,
                },
            ],
        },
    )
    assert response.status_code == 201, response.text
    return response.json()["data"]


def _upload_image(
    client: TestClient,
    product_id: int,
    image_type: str,
    *,
    sort_order: int = 0,
) -> dict[str, Any]:
    response = client.post(
        f"/api/v1/admin/products/{product_id}/images",
        files={"file": (f"{image_type}.png", _image_bytes(), "image/png")},
        data={"image_type": image_type, "sort_order": str(sort_order)},
    )
    assert response.status_code == 200, response.text
    matching = [
        image
        for image in response.json()["data"]["images"]
        if image["image_type"] == image_type and image["sort_order"] == sort_order
    ]
    assert matching
    return matching[-1]


def _has_chinese(value: str) -> bool:
    return any("\u4e00" <= character <= "\u9fff" for character in value)


def _download_template(client: TestClient) -> bytes:
    response = client.get("/api/v1/admin/products/template.xlsx")
    assert response.status_code == 200
    assert response.headers["content-type"] == XLSX_MEDIA_TYPE
    assert "harbor-market-product-template.xlsx" in response.headers["content-disposition"]
    return response.content


def _append_mapping(sheet: Any, values: dict[str, Any]) -> None:
    headers = [cell.value for cell in sheet[1]]
    sheet.append([values.get(header) for header in headers])


def _workbook_with_rows(
    template: bytes,
    *,
    products: list[dict[str, Any]],
    skus: list[dict[str, Any]] | None = None,
    images: list[dict[str, Any]] | None = None,
) -> bytes:
    workbook = load_workbook(BytesIO(template))
    for row in products:
        _append_mapping(workbook["Products"], row)
    for row in skus or []:
        _append_mapping(workbook["SKUs"], row)
    for row in images or []:
        _append_mapping(workbook["Images"], row)
    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def _product_row(
    code: str,
    *,
    category_code: str = "COFFEE",
    status: str = "draft",
    name: str = "Excel 生椰拿铁",
) -> dict[str, Any]:
    return {
        "product_code": code,
        "name": name,
        "subtitle": "批量导入",
        "category_code": category_code,
        "status": status,
        "base_price_yuan": "19.90",
        "market_price_yuan": "23.90",
        "unit": "杯",
        "stock_status": "in_stock",
        "inventory_count": 12,
        "featured": True,
        "sort_order": 3,
        "tags": "新品|咖啡",
        "selling_points": "现磨|低糖可选",
        "description": "Excel 导入商品",
        "ingredients": "咖啡、椰乳",
        "allergen_info": "含椰制品",
        "specifications_json": '[{"code":"temperature","name":"温度"}]',
    }


def _post_import(
    client: TestClient,
    payload: bytes,
    *,
    dry_run: bool,
    idempotency_key: str | None = None,
) -> Any:
    return client.post(
        "/api/v1/admin/products/import",
        params={"dry_run": str(dry_run).lower()},
        headers=({"X-Idempotency-Key": idempotency_key} if idempotency_key is not None else None),
        files={"file": ("products.xlsx", payload, XLSX_MEDIA_TYPE)},
    )


def test_excel_template_has_contract_sheets_headers_and_usage_help(
    admin_client: TestClient,
) -> None:
    _create_category(admin_client)
    workbook = load_workbook(BytesIO(_download_template(admin_client)))
    try:
        assert workbook.sheetnames == ["Products", "SKUs", "Images", "Dictionary"]
        expected_headers = {
            "Products": PRODUCT_HEADERS,
            "SKUs": SKU_HEADERS,
            "Images": IMAGE_HEADERS,
        }
        for sheet_name, headers in expected_headers.items():
            sheet = workbook[sheet_name]
            assert [cell.value for cell in sheet[1]] == headers
            assert sheet.freeze_panes == "A2"
            assert sheet.auto_filter.ref == f"A1:{sheet.cell(1, len(headers)).column_letter}1"
            assert len(sheet.data_validations.dataValidation) >= 1

        text_columns = {
            "Products": ("A", "D"),
            "SKUs": ("A", "B"),
            "Images": ("A", "C"),
        }
        for sheet_name, columns in text_columns.items():
            assert all(
                workbook[sheet_name].column_dimensions[column].number_format == "@"
                for column in columns
            )

        dictionary_values = {
            str(cell.value)
            for row in workbook["Dictionary"].iter_rows()
            for cell in row
            if cell.value is not None
        }
        assert {
            "使用流程",
            "金额",
            "product_code",
            "sku_code",
            "status",
            "stock_status",
            "image_type",
            "图片文件",
            "图片数量",
            "COFFEE",
        } <= dictionary_values
        assert any("人民币元" in value for value in dictionary_values)
        assert any("cover=1" in value for value in dictionary_values)
        assert any("唯一事实来源" in value for value in dictionary_values)
        assert any("dry_run=true" in value for value in dictionary_values)
    finally:
        workbook.close()


def test_excel_dry_run_validates_everything_without_writing_then_valid_imports(
    admin_client: TestClient,
    fake_object_storage: Any,
) -> None:
    _create_category(admin_client)
    object_key = "products/staged/CATALOG-01/0123456789abcdef0123456789abcdef.jpg"
    fake_object_storage.seed(
        object_key,
        _image_bytes(),
        content_type="image/jpeg",
    )
    workbook = _workbook_with_rows(
        _download_template(admin_client),
        products=[_product_row("CATALOG-01", status="published")],
        skus=[
            {
                "product_code": "CATALOG-01",
                "sku_code": "CATALOG-01-LARGE",
                "name": "大杯",
                "price_yuan": "21.90",
                "market_price_yuan": "25.90",
                "stock_quantity": 8,
                "attributes_json": '{"cup":"large"}',
                "is_default": True,
                "is_active": True,
                "sort_order": 1,
            }
        ],
        images=[
            {
                "product_code": "CATALOG-01",
                "image_type": "cover",
                "object_key": object_key,
                "alt_text": "Excel 商品封面",
                "sort_order": 0,
            }
        ],
    )

    dry_run = _post_import(admin_client, workbook, dry_run=True)
    assert dry_run.status_code == 200, dry_run.text
    dry_run_data = dry_run.json()["data"]
    assert {
        "dry_run": True,
        "valid": True,
        "summary": {"products": 1, "skus": 1, "images": 1, "errors": 0},
        "errors": [],
    }.items() <= dry_run_data.items()
    assert admin_client.get("/api/v1/admin/products").json()["data"]["total"] == 0
    job = admin_client.get(f"/api/v1/admin/import-jobs/{dry_run_data['job_id']}")
    assert job.status_code == 200
    assert job.json()["data"]["status"] == "validated"
    assert job.json()["data"]["dry_run"] is True
    assert set(fake_object_storage.objects) == {object_key}
    assert fake_object_storage.delete_calls == []
    assert fake_object_storage.copy_calls == []

    imported = _post_import(admin_client, workbook, dry_run=False)
    assert imported.status_code == 200, imported.text
    imported_data = imported.json()["data"]
    assert imported_data["valid"] is True
    assert imported_data["dry_run"] is False
    assert imported_data["summary"] == {
        "products": 1,
        "skus": 1,
        "images": 1,
        "errors": 0,
        "cleanup_queued": 1,
    }
    listing = admin_client.get("/api/v1/admin/products").json()["data"]
    assert listing["total"] == 1
    product = listing["items"][0]
    assert product["product_code"] == "CATALOG-01"
    assert product["base_price_cents"] == 1990
    assert product["skus"][0]["sku_code"] == "CATALOG-01-LARGE"
    permanent_key = product["images"][0]["object_key"]
    assert re.fullmatch(
        r"products/catalog/CATALOG-01/cover/[0-9a-f]{32}\.png",
        permanent_key,
    )
    assert permanent_key in fake_object_storage.objects
    assert object_key in fake_object_storage.objects
    assert fake_object_storage.copy_calls == [(object_key, permanent_key)]
    cleanup_jobs = admin_client.get("/api/v1/admin/object-cleanup-jobs").json()["data"]
    assert len(cleanup_jobs) == 2
    promotion_intent = next(
        item for item in cleanup_jobs if item["reason"] == "import_promotion_intent"
    )
    assert promotion_intent["status"] == "completed"
    staging_cleanup = next(item for item in cleanup_jobs if item["reason"] == "staging_promoted")
    assert {
        "object_key": object_key,
        "reason": "staging_promoted",
        "status": "pending",
        "attempts": 0,
        "last_error": None,
    }.items() <= staging_cleanup.items()
    retried = admin_client.post(f"/api/v1/admin/object-cleanup-jobs/{staging_cleanup['id']}/retry")
    assert retried.status_code == 200
    assert object_key not in fake_object_storage.objects
    assert admin_client.get("/api/v1/catalog/products/CATALOG-01").status_code == 200


def test_invalid_formal_excel_import_is_atomic_and_reports_row_error(
    admin_client: TestClient,
) -> None:
    _create_category(admin_client)
    workbook = _workbook_with_rows(
        _download_template(admin_client),
        products=[
            _product_row("VALID-ROW"),
            _product_row("INVALID-ROW", category_code="MISSING"),
        ],
    )

    response = _post_import(admin_client, workbook, dry_run=False)
    assert response.status_code == 200
    result = response.json()["data"]
    assert result["valid"] is False
    assert result["dry_run"] is False
    assert result["summary"] == {"products": 1, "skus": 1, "images": 0, "errors": 1}
    assert any(
        error["sheet"] == "Products" and error["row"] == 3 and error["field"] == "category_code"
        for error in result["errors"]
    )
    assert admin_client.get("/api/v1/admin/products").json()["data"]["total"] == 0
    job = admin_client.get(f"/api/v1/admin/import-jobs/{result['job_id']}").json()["data"]
    assert job["status"] == "failed"
    assert job["summary"]["errors"] == 1


def test_excel_import_idempotency_replays_same_result_and_rejects_key_reuse(
    admin_client: TestClient,
) -> None:
    _create_category(admin_client)
    template = _download_template(admin_client)
    workbook = _workbook_with_rows(
        template,
        products=[_product_row("IDEMPOTENT-01")],
    )
    key = "catalog-idempotency-test-0001"

    first = _post_import(admin_client, workbook, dry_run=True, idempotency_key=key)
    assert first.status_code == 200, first.text
    second = _post_import(admin_client, workbook, dry_run=True, idempotency_key=key)
    assert second.status_code == 200, second.text
    assert second.json()["data"] == first.json()["data"]

    changed_workbook = _workbook_with_rows(
        template,
        products=[_product_row("IDEMPOTENT-02")],
    )
    changed = _post_import(
        admin_client,
        changed_workbook,
        dry_run=True,
        idempotency_key=key,
    )
    assert changed.status_code == 409
    assert changed.json()["error"]["code"] == "idempotency_key_conflict"

    changed_mode = _post_import(
        admin_client,
        workbook,
        dry_run=False,
        idempotency_key=key,
    )
    assert changed_mode.status_code == 409
    assert changed_mode.json()["error"]["code"] == "idempotency_key_conflict"

    jobs = admin_client.get("/api/v1/admin/import-jobs").json()["data"]
    assert len(jobs) == 1
    assert jobs[0]["id"] == first.json()["data"]["job_id"]
    assert jobs[0]["idempotency_key"] == key
    assert re.fullmatch(r"[0-9a-f]{64}", jobs[0]["workbook_sha256"])


def test_pending_excel_import_idempotency_retry_reports_in_progress(
    admin_client: TestClient,
    app: FastAPI,
) -> None:
    _create_category(admin_client)
    workbook = _workbook_with_rows(
        _download_template(admin_client),
        products=[_product_row("PENDING-RETRY")],
    )
    key = "catalog-pending-retry-0001"
    with app.state.session_factory() as session:
        admin = session.scalar(select(User).where(User.username == "catalog-admin"))
        assert admin is not None
        pending = ImportJob(
            created_by=admin.id,
            status="pending",
            original_filename="products.xlsx",
            workbook_sha256=hashlib.sha256(workbook).hexdigest(),
            idempotency_key=key,
            dry_run=False,
            summary={},
            errors=[],
        )
        session.add(pending)
        session.commit()
        pending_id = pending.id

    retried = _post_import(
        admin_client,
        workbook,
        dry_run=False,
        idempotency_key=key,
    )
    assert retried.status_code == 409, retried.text
    assert retried.json()["error"]["code"] == "import_in_progress"
    jobs = admin_client.get("/api/v1/admin/import-jobs").json()["data"]
    assert [job["id"] for job in jobs] == [pending_id]
    assert jobs[0]["status"] == "pending"


def test_formal_import_completes_staging_expiry_and_queues_source_cleanup(
    admin_client: TestClient,
) -> None:
    _create_category(admin_client)
    staged_response = admin_client.post(
        "/api/v1/admin/product-images/staging",
        files={"file": ("cover.png", _image_bytes(), "image/png")},
        data={"product_code": "EXPIRY-IMPORT"},
    )
    assert staged_response.status_code == 200, staged_response.text
    source_key = staged_response.json()["data"]["object_key"]
    workbook = _workbook_with_rows(
        _download_template(admin_client),
        products=[_product_row("EXPIRY-IMPORT", status="published")],
        images=[
            {
                "product_code": "EXPIRY-IMPORT",
                "image_type": "cover",
                "object_key": source_key,
                "alt_text": "过期任务联动",
                "sort_order": 0,
            }
        ],
    )

    imported = _post_import(admin_client, workbook, dry_run=False)
    assert imported.status_code == 200, imported.text
    assert imported.json()["data"]["valid"] is True
    assert imported.json()["data"]["promoted_staging_keys"] == [source_key]
    job_id = imported.json()["data"]["job_id"]
    job = admin_client.get(f"/api/v1/admin/import-jobs/{job_id}")
    assert job.status_code == 200, job.text
    assert job.json()["data"]["promoted_staging_keys"] == [source_key]
    jobs = [
        item
        for item in admin_client.get("/api/v1/admin/object-cleanup-jobs").json()["data"]
        if item["object_key"] == source_key
    ]
    assert {item["reason"]: item["status"] for item in jobs} == {
        "staging_expiry": "completed",
        "staging_promoted": "pending",
    }


def test_export_can_be_dry_run_imported_without_key_changes(
    admin_client: TestClient,
) -> None:
    category = _create_category(admin_client)
    created = admin_client.post(
        "/api/v1/admin/products",
        json={
            "product_code": "ROUNDTRIP-01",
            "name": "往返测试拿铁",
            "category_id": category["id"],
            "base_price_cents": 2080,
            "market_price_cents": 2580,
            "description": "导出后应原样通过 dry-run",
            "tags": ["往返"],
            "selling_points": ["稳定编码"],
        },
    )
    assert created.status_code == 201, created.text
    product = created.json()["data"]
    rendered = BytesIO()
    Image.new("RGB", (2, 2), color=(10, 20, 30)).save(rendered, format="PNG")
    uploaded = admin_client.post(
        f"/api/v1/admin/products/{product['id']}/images",
        files={"file": ("roundtrip.png", rendered.getvalue(), "image/png")},
        data={"image_type": "gallery", "alt_text": "往返图片"},
    )
    assert uploaded.status_code == 200, uploaded.text

    exported = admin_client.get("/api/v1/admin/products/export.xlsx")
    assert exported.status_code == 200
    assert exported.headers["content-type"] == XLSX_MEDIA_TYPE
    assert "harbor-market-products.xlsx" in exported.headers["content-disposition"]
    workbook = load_workbook(BytesIO(exported.content), data_only=True)
    try:
        assert workbook.sheetnames == ["Products", "SKUs", "Images", "Dictionary"]
        assert workbook["Products"]["A2"].value == "ROUNDTRIP-01"
        assert workbook["Products"]["F2"].value == 20.8
        assert workbook["SKUs"]["B2"].value == "ROUNDTRIP-01-DEFAULT"
        exported_object_key = uploaded.json()["data"]["images"][0]["object_key"]
        assert workbook["Images"]["C2"].value == exported_object_key
        assert workbook["Products"]["A2"].number_format == "@"
        assert workbook["Products"]["D2"].number_format == "@"
        assert workbook["SKUs"]["A2"].number_format == "@"
        assert workbook["SKUs"]["B2"].number_format == "@"
        assert workbook["Images"]["A2"].number_format == "@"
        assert workbook["Images"]["C2"].number_format == "@"
    finally:
        workbook.close()

    roundtrip = _post_import(admin_client, exported.content, dry_run=True)
    assert roundtrip.status_code == 200, roundtrip.text
    assert roundtrip.json()["data"]["valid"] is True
    assert roundtrip.json()["data"]["summary"] == {
        "products": 1,
        "skus": 1,
        "images": 1,
        "errors": 0,
    }
    current = admin_client.get("/api/v1/admin/products").json()["data"]
    assert current["total"] == 1
    assert current["items"][0]["product_code"] == "ROUNDTRIP-01"


def test_export_recreates_deleted_catalog_records_with_canonical_keys_unchanged(
    admin_client: TestClient,
    app: FastAPI,
    fake_object_storage: Any,
) -> None:
    category = _create_category(admin_client)
    product = _create_product_with_skus(
        admin_client,
        category["id"],
        code="RESTORE-ROUNDTRIP",
    )
    direct_image = _upload_image(admin_client, product["id"], "gallery", sort_order=1)

    staged_response = admin_client.post(
        "/api/v1/admin/product-images/staging",
        files={"file": ("detail.png", _image_bytes(), "image/png")},
        data={"product_code": "RESTORE-ROUNDTRIP"},
    )
    assert staged_response.status_code == 200, staged_response.text
    promoted = _post_import(
        admin_client,
        _workbook_with_rows(
            _download_template(admin_client),
            products=[],
            images=[
                {
                    "product_code": "RESTORE-ROUNDTRIP",
                    "image_type": "detail",
                    "object_key": staged_response.json()["data"]["object_key"],
                    "alt_text": "正式目录图片",
                    "sort_order": 2,
                }
            ],
        ),
        dry_run=False,
    )
    assert promoted.status_code == 200, promoted.text
    assert promoted.json()["data"]["valid"] is True
    after_promotion = admin_client.get(f"/api/v1/admin/products/{product['id']}").json()["data"]
    catalog_image = next(
        image
        for image in after_promotion["images"]
        if image["object_key"].startswith("products/catalog/")
    )
    role_swap = _post_import(
        admin_client,
        _workbook_with_rows(
            _download_template(admin_client),
            products=[],
            images=[
                {
                    "product_code": "RESTORE-ROUNDTRIP",
                    "image_type": "cover",
                    "object_key": direct_image["object_key"],
                    "sort_order": 0,
                },
                {
                    "product_code": "RESTORE-ROUNDTRIP",
                    "image_type": "gallery",
                    "object_key": catalog_image["object_key"],
                    "sort_order": 1,
                },
            ],
        ),
        dry_run=False,
    )
    assert role_swap.status_code == 200, role_swap.text
    assert role_swap.json()["data"]["valid"] is True
    before_delete = admin_client.get(f"/api/v1/admin/products/{product['id']}").json()["data"]
    original_keys = {image["object_key"] for image in before_delete["images"]}
    original_roles = {image["object_key"]: image["image_type"] for image in before_delete["images"]}
    assert direct_image["object_key"] in original_keys
    assert any(
        re.fullmatch(
            r"products/catalog/RESTORE-ROUNDTRIP/detail/[0-9a-f]{32}\.png",
            object_key,
        )
        for object_key in original_keys
    )
    assert "/gallery/" in direct_image["object_key"]
    assert original_roles[direct_image["object_key"]] == "cover"
    assert "/detail/" in catalog_image["object_key"]
    assert original_roles[catalog_image["object_key"]] == "gallery"

    exported = admin_client.get("/api/v1/admin/products/export.xlsx")
    assert exported.status_code == 200, exported.text
    with app.state.session_factory() as session:
        stored = session.get(Product, product["id"])
        assert stored is not None
        session.delete(stored)
        session.commit()
        assert (
            session.scalar(select(Product).where(Product.product_code == "RESTORE-ROUNDTRIP"))
            is None
        )
    assert original_keys <= set(fake_object_storage.objects)
    fake_object_storage.copy_calls.clear()

    restored = _post_import(admin_client, exported.content, dry_run=False)
    assert restored.status_code == 200, restored.text
    assert restored.json()["data"]["valid"] is True
    assert restored.json()["data"]["summary"] == {
        "products": 1,
        "skus": 2,
        "images": 2,
        "errors": 0,
        "cleanup_queued": 0,
    }
    listing = admin_client.get("/api/v1/admin/products").json()["data"]
    assert listing["total"] == 1
    recreated = listing["items"][0]
    assert recreated["product_code"] == "RESTORE-ROUNDTRIP"
    assert {sku["sku_code"] for sku in recreated["skus"]} == {
        "RESTORE-ROUNDTRIP-A",
        "RESTORE-ROUNDTRIP-B",
    }
    assert {image["object_key"] for image in recreated["images"]} == original_keys
    assert {image["object_key"]: image["image_type"] for image in recreated["images"]} == (
        original_roles
    )
    assert fake_object_storage.copy_calls == []


def test_products_only_and_partial_sku_imports_preserve_unlisted_skus(
    admin_client: TestClient,
) -> None:
    category = _create_category(admin_client)
    product = _create_product_with_skus(
        admin_client,
        category["id"],
        code="PARTIAL-01",
    )
    template = _download_template(admin_client)

    products_only = _workbook_with_rows(
        template,
        products=[_product_row("PARTIAL-01", name="Products-only 更新名称")],
    )
    products_result = _post_import(admin_client, products_only, dry_run=False)
    assert products_result.status_code == 200, products_result.text
    assert products_result.json()["data"]["valid"] is True
    after_products = admin_client.get(f"/api/v1/admin/products/{product['id']}").json()["data"]
    assert after_products["name"] == "Products-only 更新名称"
    assert {sku["sku_code"] for sku in after_products["skus"]} == {
        "PARTIAL-01-A",
        "PARTIAL-01-B",
    }

    partial_sku = _workbook_with_rows(
        template,
        products=[],
        skus=[
            {
                "product_code": "PARTIAL-01",
                "sku_code": "PARTIAL-01-B",
                "name": "规格 B（Excel 更新）",
                "price_yuan": "22.90",
                "market_price_yuan": "26.90",
                "stock_quantity": 3,
                "attributes_json": '{"cup":"large"}',
                "is_default": False,
                "is_active": True,
                "sort_order": 2,
            }
        ],
    )
    sku_result = _post_import(admin_client, partial_sku, dry_run=False)
    assert sku_result.status_code == 200, sku_result.text
    assert sku_result.json()["data"]["valid"] is True
    after_sku = admin_client.get(f"/api/v1/admin/products/{product['id']}").json()["data"]
    by_code = {sku["sku_code"]: sku for sku in after_sku["skus"]}
    assert set(by_code) == {"PARTIAL-01-A", "PARTIAL-01-B"}
    assert by_code["PARTIAL-01-A"]["price_cents"] == 1990
    assert by_code["PARTIAL-01-A"]["is_default"] is True
    assert by_code["PARTIAL-01-B"]["name"] == "规格 B（Excel 更新）"
    assert by_code["PARTIAL-01-B"]["price_cents"] == 2290


def test_orphan_sku_reports_chinese_error_at_source_row(
    admin_client: TestClient,
) -> None:
    _create_category(admin_client)
    workbook = _workbook_with_rows(
        _download_template(admin_client),
        products=[],
        skus=[
            {
                "product_code": "MISSING-PRODUCT",
                "sku_code": "ORPHAN-SKU",
                "name": "孤立规格",
                "price_yuan": "9.90",
                "stock_quantity": 1,
                "is_default": True,
                "is_active": True,
            }
        ],
    )

    response = _post_import(admin_client, workbook, dry_run=True)
    assert response.status_code == 200
    result = response.json()["data"]
    assert result["valid"] is False
    matching = [
        error
        for error in result["errors"]
        if error["sheet"] == "SKUs" and error["row"] == 2 and error["field"] == "product_code"
    ]
    assert len(matching) == 1
    assert _has_chinese(matching[0]["message"])
    assert "不存在" in matching[0]["message"]


def test_images_only_import_can_move_existing_image_role(
    admin_client: TestClient,
) -> None:
    category = _create_category(admin_client)
    product = _create_product_with_skus(admin_client, category["id"], code="IMAGE-MOVE")
    gallery = _upload_image(admin_client, product["id"], "gallery")
    workbook = _workbook_with_rows(
        _download_template(admin_client),
        products=[],
        images=[
            {
                "product_code": "IMAGE-MOVE",
                "image_type": "cover",
                "object_key": gallery["object_key"],
                "alt_text": "改为封面",
                "sort_order": 0,
            }
        ],
    )

    response = _post_import(admin_client, workbook, dry_run=False)
    assert response.status_code == 200, response.text
    assert response.json()["data"]["valid"] is True
    current = admin_client.get(f"/api/v1/admin/products/{product['id']}").json()["data"]
    assert len(current["images"]) == 1
    assert current["images"][0]["object_key"] == gallery["object_key"]
    assert current["images"][0]["image_type"] == "cover"


@pytest.mark.parametrize(
    "upload_order",
    [("cover", "gallery"), ("gallery", "cover")],
)
def test_images_only_import_can_swap_cover_in_either_database_id_direction(
    admin_client: TestClient,
    upload_order: tuple[str, str],
) -> None:
    category = _create_category(admin_client)
    product = _create_product_with_skus(admin_client, category["id"], code="IMAGE-SWAP")
    uploaded = {
        image_type: _upload_image(admin_client, product["id"], image_type)
        for image_type in upload_order
    }
    old_cover = uploaded["cover"]
    old_gallery = uploaded["gallery"]
    if upload_order[0] == "cover":
        assert old_cover["id"] < old_gallery["id"]
    else:
        assert old_gallery["id"] < old_cover["id"]

    workbook = _workbook_with_rows(
        _download_template(admin_client),
        products=[],
        images=[
            {
                "product_code": "IMAGE-SWAP",
                "image_type": "gallery",
                "object_key": old_cover["object_key"],
                "alt_text": "原封面改轮播",
                "sort_order": 1,
            },
            {
                "product_code": "IMAGE-SWAP",
                "image_type": "cover",
                "object_key": old_gallery["object_key"],
                "alt_text": "原轮播改封面",
                "sort_order": 0,
            },
        ],
    )

    response = _post_import(admin_client, workbook, dry_run=False)
    assert response.status_code == 200, response.text
    assert response.json()["data"]["valid"] is True
    current = admin_client.get(f"/api/v1/admin/products/{product['id']}").json()["data"]
    by_key = {image["object_key"]: image for image in current["images"]}
    assert by_key[old_cover["object_key"]]["id"] == old_cover["id"]
    assert by_key[old_cover["object_key"]]["image_type"] == "gallery"
    assert by_key[old_gallery["object_key"]]["id"] == old_gallery["id"]
    assert by_key[old_gallery["object_key"]]["image_type"] == "cover"
    assert sum(image["image_type"] == "cover" for image in current["images"]) == 1


def test_images_only_role_moves_enforce_final_cover_and_gallery_limits(
    admin_client: TestClient,
) -> None:
    category = _create_category(admin_client)
    product = _create_product_with_skus(admin_client, category["id"], code="IMAGE-LIMIT")
    cover = _upload_image(admin_client, product["id"], "cover")
    galleries = [
        _upload_image(admin_client, product["id"], "gallery", sort_order=index)
        for index in range(8)
    ]
    template = _download_template(admin_client)

    second_cover = _workbook_with_rows(
        template,
        products=[],
        images=[
            {
                "product_code": "IMAGE-LIMIT",
                "image_type": "cover",
                "object_key": galleries[0]["object_key"],
                "sort_order": 1,
            }
        ],
    )
    cover_result = _post_import(admin_client, second_cover, dry_run=True).json()["data"]
    assert cover_result["valid"] is False
    assert any(
        "cover" in error["message"] and "1" in error["message"] for error in cover_result["errors"]
    )

    ninth_gallery = _workbook_with_rows(
        template,
        products=[],
        images=[
            {
                "product_code": "IMAGE-LIMIT",
                "image_type": "gallery",
                "object_key": cover["object_key"],
                "sort_order": 9,
            }
        ],
    )
    gallery_result = _post_import(admin_client, ninth_gallery, dry_run=True).json()["data"]
    assert gallery_result["valid"] is False
    assert any(
        "gallery" in error["message"] and "8" in error["message"]
        for error in gallery_result["errors"]
    )

    unchanged = admin_client.get(f"/api/v1/admin/products/{product['id']}").json()["data"]
    assert sum(image["image_type"] == "cover" for image in unchanged["images"]) == 1
    assert sum(image["image_type"] == "gallery" for image in unchanged["images"]) == 8


def test_excel_rejects_invalid_staged_content_and_wrong_canonical_binding(
    admin_client: TestClient,
    fake_object_storage: Any,
) -> None:
    category = _create_category(admin_client)
    _create_product_with_skus(admin_client, category["id"], code="OBJECT-CHECK")
    template = _download_template(admin_client)
    cases = [
        (
            "products/staged/OBJECT-CHECK/11111111111111111111111111111111.png",
            b"not really a png",
            "image/png",
            None,
        ),
        (
            "products/staged/OBJECT-CHECK/22222222222222222222222222222222.png",
            _image_bytes(),
            "image/png",
            5 * 1024 * 1024 + 1,
        ),
        (
            "products/catalog/OTHER-PRODUCT/gallery/33333333333333333333333333333333.png",
            _image_bytes(),
            "image/png",
            None,
        ),
    ]

    for index, (object_key, payload, content_type, reported_size) in enumerate(cases):
        fake_object_storage.seed(
            object_key,
            payload,
            content_type=content_type,
            reported_size=reported_size,
        )
        workbook = _workbook_with_rows(
            template,
            products=[],
            images=[
                {
                    "product_code": "OBJECT-CHECK",
                    "image_type": "gallery",
                    "object_key": object_key,
                    "sort_order": index,
                }
            ],
        )
        response = _post_import(admin_client, workbook, dry_run=True)
        assert response.status_code == 200
        result = response.json()["data"]
        assert result["valid"] is False, (object_key, result)
        matching = [
            error
            for error in result["errors"]
            if error["sheet"] == "Images" and error["row"] == 2 and error["field"] == "object_key"
        ]
        assert matching, (object_key, result["errors"])
        assert _has_chinese(matching[0]["message"])


def test_excel_rejects_cross_product_canonical_key_reassignment(
    admin_client: TestClient,
) -> None:
    category = _create_category(admin_client)
    source = _create_product_with_skus(admin_client, category["id"], code="KEY-OWNER")
    target = _create_product_with_skus(admin_client, category["id"], code="KEY-TARGET")
    image = _upload_image(admin_client, source["id"], "gallery")
    workbook = _workbook_with_rows(
        _download_template(admin_client),
        products=[],
        images=[
            {
                "product_code": "KEY-TARGET",
                "image_type": "gallery",
                "object_key": image["object_key"],
            }
        ],
    )

    response = _post_import(admin_client, workbook, dry_run=True)
    assert response.status_code == 200, response.text
    result = response.json()["data"]
    assert result["valid"] is False
    assert any(
        error["sheet"] == "Images"
        and error["field"] == "object_key"
        and "其他商品" in error["message"]
        for error in result["errors"]
    )
    unchanged = admin_client.get(f"/api/v1/admin/products/{source['id']}").json()["data"]
    assert [item["object_key"] for item in unchanged["images"]] == [image["object_key"]]
    assert admin_client.get(f"/api/v1/admin/products/{target['id']}").json()["data"]["images"] == []


@pytest.mark.parametrize("price", ["NaN", "Infinity", "-Infinity"])
def test_excel_rejects_non_finite_prices(
    admin_client: TestClient,
    price: str,
) -> None:
    _create_category(admin_client)
    product = _product_row("NON-FINITE")
    product["base_price_yuan"] = price
    workbook = _workbook_with_rows(
        _download_template(admin_client),
        products=[product],
    )

    response = _post_import(admin_client, workbook, dry_run=True)
    assert response.status_code == 200
    result = response.json()["data"]
    assert result["valid"] is False
    assert any(
        error["sheet"] == "Products"
        and error["row"] == 2
        and error["field"] == "base_price_yuan"
        and _has_chinese(error["message"])
        and "价格" in error["message"]
        for error in result["errors"]
    )
    assert admin_client.get("/api/v1/admin/products").json()["data"]["total"] == 0


def test_excel_reports_pydantic_and_parser_errors_at_actual_columns(
    admin_client: TestClient,
) -> None:
    _create_category(admin_client)
    missing_name = _product_row("FIELD-PRODUCT", name="")
    workbook = _workbook_with_rows(
        _download_template(admin_client),
        products=[missing_name],
        skus=[
            {
                "product_code": "FIELD-PRODUCT",
                "sku_code": "FIELD-PRODUCT-LARGE",
                "name": "大杯",
                "price_yuan": "not-a-price",
                "stock_quantity": 1,
                "attributes_json": "{}",
                "is_default": True,
                "is_active": True,
                "sort_order": 0,
            }
        ],
    )

    result = _post_import(admin_client, workbook, dry_run=True).json()["data"]

    assert result["valid"] is False
    assert {
        (error["sheet"], error["row"], error["field"])
        for error in result["errors"]
    } >= {
        ("Products", 2, "name"),
        ("SKUs", 2, "price_yuan"),
    }


@pytest.mark.parametrize(
    ("skus", "expected_row", "expected_field"),
    [
        (
            [
                {
                    "product_code": "SKU-MERGE-FIELD",
                    "sku_code": "SKU-MERGE-FIELD-A",
                    "name": "规格 A",
                    "price_yuan": "19.90",
                    "is_default": True,
                    "is_active": True,
                },
                {
                    "product_code": "SKU-MERGE-FIELD",
                    "sku_code": "SKU-MERGE-FIELD-B",
                    "name": "规格 B",
                    "price_yuan": "20.90",
                    "is_default": True,
                    "is_active": True,
                },
            ],
            3,
            "is_default",
        ),
        (
            [
                {
                    "product_code": "SKU-MERGE-FIELD",
                    "sku_code": "SKU-MERGE-FIELD-INACTIVE",
                    "name": "停用规格",
                    "price_yuan": "19.90",
                    "is_default": False,
                    "is_active": False,
                }
            ],
            2,
            "is_active",
        ),
    ],
)
def test_excel_reports_final_sku_errors_on_sku_source_row(
    admin_client: TestClient,
    skus: list[dict[str, Any]],
    expected_row: int,
    expected_field: str,
) -> None:
    _create_category(admin_client)
    workbook = _workbook_with_rows(
        _download_template(admin_client),
        products=[_product_row("SKU-MERGE-FIELD")],
        skus=skus,
    )

    result = _post_import(admin_client, workbook, dry_run=True).json()["data"]

    assert result["valid"] is False
    assert any(
        error["sheet"] == "SKUs"
        and error["row"] == expected_row
        and error["field"] == expected_field
        for error in result["errors"]
    )
    assert not any(
        error["sheet"] == "Products" and error["field"] == expected_field
        for error in result["errors"]
    )


def test_excel_maps_product_and_sku_price_comparisons_to_market_price_column(
    admin_client: TestClient,
) -> None:
    _create_category(admin_client)
    product = _product_row("MARKET-COMPARISON")
    product["base_price_yuan"] = "20.00"
    product["market_price_yuan"] = "19.00"
    workbook = _workbook_with_rows(
        _download_template(admin_client),
        products=[product],
        skus=[
            {
                "product_code": "MARKET-COMPARISON",
                "sku_code": "MARKET-COMPARISON-LARGE",
                "name": "大杯",
                "price_yuan": "20.00",
                "market_price_yuan": "19.00",
                "stock_quantity": 1,
                "is_default": True,
                "is_active": True,
            }
        ],
    )

    result = _post_import(admin_client, workbook, dry_run=True).json()["data"]

    assert result["valid"] is False
    assert {
        (error["sheet"], error["row"], error["field"])
        for error in result["errors"]
    } >= {
        ("Products", 2, "market_price_yuan"),
        ("SKUs", 2, "market_price_yuan"),
    }


def test_excel_rejects_image_alt_text_longer_than_200_characters(
    admin_client: TestClient,
) -> None:
    category = _create_category(admin_client)
    product = _create_product_with_skus(admin_client, category["id"], code="ALT-LIMIT")
    image = _upload_image(admin_client, product["id"], "gallery")
    workbook = _workbook_with_rows(
        _download_template(admin_client),
        products=[],
        images=[
            {
                "product_code": "ALT-LIMIT",
                "image_type": "gallery",
                "object_key": image["object_key"],
                "alt_text": "图" * 201,
                "sort_order": 0,
            }
        ],
    )

    result = _post_import(admin_client, workbook, dry_run=True).json()["data"]
    assert result["valid"] is False
    assert any(
        error["sheet"] == "Images"
        and error["row"] == 2
        and error["field"] == "alt_text"
        and "200" in error["message"]
        for error in result["errors"]
    )


def test_excel_rejects_excessive_zip_compression_ratio(
    admin_client: TestClient,
) -> None:
    payload = BytesIO()
    with zipfile.ZipFile(
        payload,
        mode="w",
        compression=zipfile.ZIP_DEFLATED,
        compresslevel=9,
    ) as archive:
        archive.writestr("xl/worksheets/sheet1.xml", b"0" * (8 * 1024 * 1024))

    response = _post_import(admin_client, payload.getvalue(), dry_run=True)
    assert response.status_code == 200
    result = response.json()["data"]
    assert result["valid"] is False
    assert any(
        error["sheet"] == "Workbook" and error["field"] == "file" and "压缩比" in error["message"]
        for error in result["errors"]
    )


def test_excel_image_row_limit_accepts_boundary_then_rejects_next_row(
    admin_client: TestClient,
    fake_object_storage: Any,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    category = _create_category(admin_client)
    product = _create_product_with_skus(admin_client, category["id"], code="ROW-BOUNDARY")
    monkeypatch.setattr(catalog_excel, "_MAX_IMPORT_IMAGE_ROWS", 2)
    image_rows = []
    for index in range(3):
        object_key = f"products/staged/ROW-BOUNDARY/{index:032x}.png"
        if index < 2:
            fake_object_storage.seed(object_key, _image_bytes(), content_type="image/png")
        image_rows.append(
            {
                "product_code": "ROW-BOUNDARY",
                "image_type": "detail",
                "object_key": object_key,
                "sort_order": index,
            }
        )
    workbook = _workbook_with_rows(
        _download_template(admin_client),
        products=[],
        images=image_rows,
    )

    result = _post_import(admin_client, workbook, dry_run=True).json()["data"]
    assert result["valid"] is False
    assert result["summary"]["images"] == 2
    assert any(
        error["sheet"] == "Images"
        and error["row"] == 4
        and error["field"] == "rows"
        and "2" in error["message"]
        for error in result["errors"]
    )
    assert admin_client.get(f"/api/v1/admin/products/{product['id']}").status_code == 200


def test_excel_total_image_bytes_accepts_boundary_then_rejects_next_image(
    admin_client: TestClient,
    fake_object_storage: Any,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    category = _create_category(admin_client)
    _create_product_with_skus(admin_client, category["id"], code="BYTE-BOUNDARY")
    payload = _image_bytes()
    monkeypatch.setattr(catalog_excel, "_MAX_IMPORT_IMAGE_BYTES", len(payload) * 2)
    image_rows = []
    for index in range(3):
        object_key = f"products/staged/BYTE-BOUNDARY/{index + 10:032x}.png"
        fake_object_storage.seed(object_key, payload, content_type="image/png")
        image_rows.append(
            {
                "product_code": "BYTE-BOUNDARY",
                "image_type": "detail",
                "object_key": object_key,
                "sort_order": index,
            }
        )
    workbook = _workbook_with_rows(
        _download_template(admin_client),
        products=[],
        images=image_rows,
    )

    result = _post_import(admin_client, workbook, dry_run=True).json()["data"]
    assert result["valid"] is False
    assert result["summary"]["images"] == 2
    assert any(
        error["sheet"] == "Images"
        and error["row"] == 4
        and error["field"] == "object_key"
        and "总大小" in error["message"]
        for error in result["errors"]
    )


@pytest.mark.parametrize("failure_mode", ["copy_after_destination", "database_flush"])
def test_failed_formal_import_records_retryable_destination_cleanup(
    admin_client: TestClient,
    fake_object_storage: Any,
    app: FastAPI,
    failure_mode: str,
) -> None:
    category = _create_category(admin_client)
    _create_product_with_skus(admin_client, category["id"], code="ROLLBACK-IMAGE")
    source_key = "products/staged/ROLLBACK-IMAGE/abcdefabcdefabcdefabcdefabcdefab.jpg"
    fake_object_storage.seed(source_key, _image_bytes(), content_type="image/jpeg")
    workbook = _workbook_with_rows(
        _download_template(admin_client),
        products=[],
        images=[
            {
                "product_code": "ROLLBACK-IMAGE",
                "image_type": "gallery",
                "object_key": source_key,
                "alt_text": "等待回滚的图片",
                "sort_order": 0,
            }
        ],
    )
    if failure_mode == "copy_after_destination":
        fake_object_storage.copy_then_fail = True
    else:
        with app.state.engine.begin() as connection:
            connection.exec_driver_sql(
                """
                CREATE TRIGGER reject_imported_image
                BEFORE INSERT ON product_images
                BEGIN
                    SELECT RAISE(ABORT, 'forced image insert failure');
                END
                """
            )
    fake_object_storage.fail_delete = True

    response = _post_import(admin_client, workbook, dry_run=False)
    assert response.status_code == 200, response.text
    result = response.json()["data"]
    assert result["valid"] is False
    job = admin_client.get(f"/api/v1/admin/import-jobs/{result['job_id']}").json()["data"]
    assert job["status"] == "failed"

    pending_cleanup = admin_client.get(
        "/api/v1/admin/object-cleanup-jobs", params={"status": "pending"}
    ).json()["data"]
    assert len(pending_cleanup) == 1
    cleanup = pending_cleanup[0]
    destination_key = cleanup["object_key"]
    assert cleanup["reason"] == "import_promotion_intent"
    assert cleanup["attempts"] == 0
    assert re.fullmatch(
        r"products/catalog/ROLLBACK-IMAGE/gallery/[0-9a-f]{32}\.png",
        destination_key,
    )
    assert destination_key in fake_object_storage.objects
    assert source_key in fake_object_storage.objects

    first_retry = admin_client.post(f"/api/v1/admin/object-cleanup-jobs/{cleanup['id']}/retry")
    assert first_retry.status_code == 503
    cleanup = admin_client.get(
        "/api/v1/admin/object-cleanup-jobs", params={"status": "failed"}
    ).json()["data"][0]
    assert cleanup["attempts"] == 1

    fake_object_storage.fail_delete = False
    fake_object_storage.copy_then_fail = False
    retried = admin_client.post(f"/api/v1/admin/object-cleanup-jobs/{cleanup['id']}/retry")
    assert retried.status_code == 200, retried.text
    assert retried.json()["data"]["status"] == "completed"
    assert retried.json()["data"]["attempts"] == 2
    assert destination_key not in fake_object_storage.objects
    assert source_key in fake_object_storage.objects
