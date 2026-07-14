from __future__ import annotations

from datetime import UTC, datetime
from io import BytesIO
from typing import Any

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from pydantic import BaseModel, ValidationError

from app.schemas.catalog import (
    MAX_PRODUCT_SPECIFICATIONS,
    MAX_SKU_ATTRIBUTES,
    MAX_SKU_ATTRIBUTES_JSON_CHARACTERS,
    MAX_SPECIFICATION_OPTIONS,
    MAX_SPECIFICATIONS_JSON_CHARACTERS,
    POSTGRES_INTEGER_MAX,
    POSTGRES_INTEGER_MIN,
    CategoryCreate,
    CategoryUpdate,
    ImageUpdate,
    ProductCreate,
    ProductRead,
    ProductSkuInput,
    ProductSpecification,
    ProductUpdate,
    SpecificationOption,
)


def _product_payload(**overrides: Any) -> dict[str, Any]:
    payload: dict[str, Any] = {
        "product_code": "COFFEE-1",
        "name": "Coffee",
        "category_id": 1,
        "base_price_cents": 1_990,
    }
    payload.update(overrides)
    return payload


def _size_specification() -> dict[str, Any]:
    return {
        "code": "size",
        "name": "Size",
        "selection_mode": "single",
        "required": True,
        "min_select": 1,
        "max_select": 1,
        "options": [
            {
                "code": "small",
                "name": "Small",
                "price_delta_cents": 0,
                "sort": 0,
                "is_default": True,
            },
            {
                "code": "large",
                "name": "Large",
                "price_delta_cents": 300,
                "sort": 1,
                "is_default": False,
            },
        ],
    }


def _product_read_payload(specifications: list[dict[str, Any]]) -> dict[str, Any]:
    timestamp = datetime.now(UTC)
    return {
        "id": 1,
        "product_code": "COFFEE-1",
        "name": "Coffee",
        "subtitle": None,
        "category": {
            "id": 1,
            "code": "COFFEE",
            "name": "Coffee",
            "description": None,
            "parent_id": None,
            "sort_order": 0,
            "is_active": True,
            "created_at": timestamp,
            "updated_at": timestamp,
        },
        "status": "draft",
        "base_price_cents": 1_990,
        "market_price_cents": None,
        "currency": "CNY",
        "unit": "杯",
        "description": "",
        "featured": False,
        "stock_status": "in_stock",
        "inventory_count": None,
        "tags": [],
        "selling_points": [],
        "specifications": specifications,
        "ingredients": None,
        "allergen_info": None,
        "sort_order": 0,
        "skus": [],
        "images": [],
        "created_at": timestamp,
        "updated_at": timestamp,
    }


def test_legacy_code_name_specification_is_compatible_across_write_and_read_models() -> None:
    legacy = {"code": "temperature", "name": "Temperature"}

    created = ProductCreate.model_validate(_product_payload(specifications=[legacy]))
    updated = ProductUpdate.model_validate({"specifications": [legacy]})
    read = ProductRead.model_validate(_product_read_payload([legacy]))
    assert updated.specifications is not None

    for specification in (
        created.specifications[0],
        updated.specifications[0],
        read.specifications[0],
    ):
        assert isinstance(specification, ProductSpecification)
        assert specification.selection_mode == "single"
        assert specification.required is False
        assert specification.min_select == 0
        assert specification.max_select == 1
        assert specification.options == []


def test_complete_specification_and_sku_reference_are_normalized() -> None:
    product = ProductCreate.model_validate(
        _product_payload(
            specifications=[_size_specification()],
            skus=[
                {
                    "sku_code": "COFFEE-1-LARGE",
                    "name": "Large coffee",
                    "price_cents": 2_290,
                    "attributes": {"size": "large"},
                }
            ],
        )
    )

    assert product.specifications[0].options[1] == SpecificationOption(
        code="large",
        name="Large",
        price_delta_cents=300,
        sort=1,
    )
    assert product.skus[0].attributes == {"size": "large"}


@pytest.mark.parametrize("code", ["", "bad code", "bad/code", "x" * 65, "bad\x00code"])
def test_specification_codes_are_bounded_and_safe(code: str) -> None:
    with pytest.raises(ValidationError):
        ProductSpecification(code=code, name="Size")
    with pytest.raises(ValidationError):
        SpecificationOption(code=code, name="Large")
    with pytest.raises(ValidationError):
        ProductSkuInput(
            sku_code="SKU-1",
            name="SKU",
            price_cents=0,
            attributes={code: "large"},
        )


def test_specification_option_and_attribute_counts_are_bounded() -> None:
    specifications = [
        {"code": f"spec-{index}", "name": f"Spec {index}"}
        for index in range(MAX_PRODUCT_SPECIFICATIONS + 1)
    ]
    options = [
        {"code": f"option-{index}", "name": f"Option {index}"}
        for index in range(MAX_SPECIFICATION_OPTIONS + 1)
    ]
    attributes = {f"spec-{index}": f"option-{index}" for index in range(MAX_SKU_ATTRIBUTES + 1)}

    with pytest.raises(ValidationError):
        ProductCreate.model_validate(_product_payload(specifications=specifications))
    with pytest.raises(ValidationError):
        ProductSpecification(code="size", name="Size", options=options)
    with pytest.raises(ValidationError):
        ProductSkuInput(sku_code="SKU-1", name="SKU", price_cents=0, attributes=attributes)


def test_specification_and_option_codes_must_be_unique() -> None:
    with pytest.raises(ValidationError, match="specification codes must be unique"):
        ProductCreate.model_validate(
            _product_payload(
                specifications=[
                    {"code": "size", "name": "Size"},
                    {"code": "SIZE", "name": "Duplicate size"},
                ]
            )
        )

    with pytest.raises(ValidationError, match="option codes must be unique"):
        ProductSpecification(
            code="size",
            name="Size",
            options=[
                {"code": "large", "name": "Large"},
                {"code": "LARGE", "name": "Duplicate large"},
            ],
        )

    with pytest.raises(ValidationError, match="attribute specification codes must be unique"):
        ProductSkuInput(
            sku_code="SKU-1",
            name="SKU",
            price_cents=0,
            attributes={"size": "small", "SIZE": "large"},
        )


@pytest.mark.parametrize(
    "overrides",
    [
        {"max_select": 0},
        {"required": True, "min_select": 0},
        {"required": False, "min_select": 1},
        {"selection_mode": "multiple", "required": True, "min_select": 2, "max_select": 1},
        {"selection_mode": "single", "max_select": 2},
        {"selection_mode": "multiple", "required": True, "min_select": 3, "max_select": 3},
        {
            "options": [
                {"code": "small", "name": "Small", "is_default": True},
                {"code": "large", "name": "Large", "is_default": True},
            ]
        },
        {
            "selection_mode": "multiple",
            "max_select": 1,
            "options": [
                {"code": "small", "name": "Small", "is_default": True},
                {"code": "large", "name": "Large", "is_default": True},
            ],
        },
    ],
)
def test_selection_contract_rejects_inconsistent_bounds_and_defaults(
    overrides: dict[str, Any],
) -> None:
    payload: dict[str, Any] = {
        "code": "size",
        "name": "Size",
        "options": [
            {"code": "small", "name": "Small"},
            {"code": "large", "name": "Large"},
        ],
    }
    payload.update(overrides)

    with pytest.raises(ValidationError):
        ProductSpecification.model_validate(payload)


def test_option_integer_fields_accept_only_signed_int32_values() -> None:
    minimum = SpecificationOption(
        code="discount",
        name="Discount",
        price_delta_cents=POSTGRES_INTEGER_MIN,
        sort=POSTGRES_INTEGER_MIN,
    )
    maximum = SpecificationOption(
        code="premium",
        name="Premium",
        price_delta_cents=POSTGRES_INTEGER_MAX,
        sort=POSTGRES_INTEGER_MAX,
    )
    assert minimum.price_delta_cents == POSTGRES_INTEGER_MIN
    assert maximum.sort == POSTGRES_INTEGER_MAX

    for field_name, value in (
        ("price_delta_cents", POSTGRES_INTEGER_MIN - 1),
        ("price_delta_cents", POSTGRES_INTEGER_MAX + 1),
        ("sort", POSTGRES_INTEGER_MIN - 1),
        ("sort", POSTGRES_INTEGER_MAX + 1),
    ):
        with pytest.raises(ValidationError):
            SpecificationOption.model_validate(
                {"code": "option", "name": "Option", field_name: value}
            )


@pytest.mark.parametrize("value", [float("nan"), float("inf"), float("-inf")])
def test_non_finite_json_numbers_are_rejected(value: float) -> None:
    with pytest.raises(ValidationError):
        SpecificationOption(
            code="option",
            name="Option",
            price_delta_cents=value,
        )
    with pytest.raises(ValidationError):
        ProductSkuInput(
            sku_code="SKU-1",
            name="SKU",
            price_cents=0,
            attributes={"size": value},
        )


@pytest.mark.parametrize("control", ["\x00", "\x01", "\x08", "\x0b", "\x0c", "\x0e", "\x1f"])
def test_xml_illegal_c0_control_characters_are_rejected(control: str) -> None:
    with pytest.raises(ValidationError, match="XML-illegal C0"):
        ProductSpecification(code="size", name=f"Si{control}ze")
    with pytest.raises(ValidationError, match="XML-illegal C0"):
        SpecificationOption(code="large", name=f"Lar{control}ge")
    with pytest.raises(ValidationError, match="XML-illegal C0"):
        ProductSpecification(code=f"{control}size", name="Size")
    with pytest.raises(ValidationError, match="XML-illegal C0"):
        ProductSkuInput(
            sku_code="SKU-1",
            name="SKU",
            price_cents=0,
            attributes={f"{control}size": "large"},
        )


@pytest.mark.parametrize(
    ("model_type", "payload"),
    [
        (
            CategoryCreate,
            {"code": "COFFEE", "name": "Coffee", "description": "bad\x00description"},
        ),
        (CategoryUpdate, {"description": "bad\x0bdescription"}),
        (ProductCreate, _product_payload(subtitle="bad\x0csubtitle")),
        (ProductCreate, _product_payload(tags=["safe", "bad\x1ftag"])),
        (
            ProductCreate,
            _product_payload(
                specifications=[
                    {
                        "code": "size",
                        "name": "Size",
                        "options": [{"code": "large", "name": "bad\x00option"}],
                    }
                ]
            ),
        ),
        (
            ProductSkuInput,
            {"sku_code": "SKU-1", "name": "bad\x0bsku", "price_cents": 0},
        ),
        (
            ProductSkuInput,
            {
                "sku_code": "SKU-1",
                "name": "SKU",
                "price_cents": 0,
                "attributes": {"size": "bad\x0coption"},
            },
        ),
        (ImageUpdate, {"alt_text": "bad\x00alt text"}),
    ],
)
def test_strict_models_reject_control_characters_in_persisted_text_and_json(
    model_type: type[BaseModel],
    payload: dict[str, Any],
) -> None:
    with pytest.raises(ValidationError, match="XML-illegal C0"):
        model_type.model_validate(payload)


def test_strict_models_allow_xml_legal_tab_lf_and_cr() -> None:
    category = CategoryCreate(
        code="COFFEE",
        name="Coffee",
        description="line one\tvalue\nline two\rline three",
    )

    assert category.description == "line one\tvalue\nline two\rline three"


def test_arbitrary_nested_sku_attribute_json_is_rejected() -> None:
    with pytest.raises(ValidationError):
        ProductSkuInput(
            sku_code="SKU-1",
            name="SKU",
            price_cents=0,
            attributes={"size": {"code": "large"}},
        )


@pytest.mark.parametrize(
    ("specifications", "attributes", "message"),
    [
        ([_size_specification()], {"temperature": "hot"}, "unknown specification"),
        ([_size_specification()], {"size": "extra-large"}, "unknown option"),
        ([], {"size": "large"}, "unknown specification"),
    ],
)
def test_product_create_rejects_unknown_sku_attribute_references(
    specifications: list[dict[str, Any]],
    attributes: dict[str, str],
    message: str,
) -> None:
    with pytest.raises(ValidationError, match=message):
        ProductCreate.model_validate(
            _product_payload(
                specifications=specifications,
                skus=[
                    {
                        "sku_code": "SKU-1",
                        "name": "SKU",
                        "price_cents": 0,
                        "attributes": attributes,
                    }
                ],
            )
        )


def test_product_update_validates_references_only_when_both_sides_are_available() -> None:
    sku = {
        "sku_code": "SKU-1",
        "name": "SKU",
        "price_cents": 0,
        "attributes": {"size": "unknown"},
    }

    ProductUpdate.model_validate({"skus": [sku]})
    ProductUpdate.model_validate({"specifications": [_size_specification()]})
    with pytest.raises(ValidationError, match="unknown option"):
        ProductUpdate.model_validate({"specifications": [_size_specification()], "skus": [sku]})


def test_legacy_specs_without_options_preserve_unverifiable_sku_attributes() -> None:
    product = ProductCreate.model_validate(
        _product_payload(
            specifications=[{"code": "temperature", "name": "Temperature"}],
            skus=[
                {
                    "sku_code": "SKU-1",
                    "name": "SKU",
                    "price_cents": 0,
                    "attributes": {"legacy-size": "large"},
                }
            ],
        )
    )

    assert product.skus[0].attributes == {"legacy-size": "large"}


def test_specification_and_attribute_json_have_safe_serialized_size_limits() -> None:
    assert MAX_SPECIFICATIONS_JSON_CHARACTERS < 32_767
    assert MAX_SKU_ATTRIBUTES_JSON_CHARACTERS < MAX_SPECIFICATIONS_JSON_CHARACTERS

    oversized_specifications = [
        {
            "code": f"spec-{specification_index}",
            "name": "S" * 120,
            "selection_mode": "multiple",
            "max_select": MAX_SPECIFICATION_OPTIONS,
            "options": [
                {
                    "code": f"option-{option_index}",
                    "name": "O" * 120,
                }
                for option_index in range(MAX_SPECIFICATION_OPTIONS)
            ],
        }
        for specification_index in range(MAX_PRODUCT_SPECIFICATIONS)
    ]
    oversized_attributes = {
        f"k{index:02d}" + "k" * 61: f"v{index:02d}" + "v" * 61
        for index in range(MAX_SKU_ATTRIBUTES)
    }

    with pytest.raises(ValidationError, match="serialize to at most"):
        ProductCreate.model_validate(_product_payload(specifications=oversized_specifications))
    with pytest.raises(ValidationError, match="serialize to at most"):
        ProductSkuInput(
            sku_code="SKU-1",
            name="SKU",
            price_cents=0,
            attributes=oversized_attributes,
        )


def test_api_rejects_illegal_characters_before_safe_excel_export(
    admin_client: TestClient,
) -> None:
    rejected_category = admin_client.post(
        "/api/v1/admin/categories",
        json={
            "code": "REJECTED",
            "name": "Rejected",
            "description": "illegal\x00description",
        },
    )
    assert rejected_category.status_code == 422
    assert rejected_category.json()["error"]["code"] == "validation_error"

    category_response = admin_client.post(
        "/api/v1/admin/categories",
        json={"code": "EXPORT", "name": "Export-safe category"},
    )
    assert category_response.status_code == 201
    category_id = category_response.json()["data"]["id"]

    rejected_product = admin_client.post(
        "/api/v1/admin/products",
        json=_product_payload(
            product_code="REJECTED-PRODUCT",
            category_id=category_id,
            tags=["safe", "illegal\x0btag"],
        ),
    )
    assert rejected_product.status_code == 422
    assert rejected_product.json()["error"]["code"] == "validation_error"

    created_product = admin_client.post(
        "/api/v1/admin/products",
        json=_product_payload(
            product_code="SAFE-EXPORT",
            category_id=category_id,
            specifications=[_size_specification()],
            skus=[
                {
                    "sku_code": "SAFE-EXPORT-SMALL",
                    "name": "Small",
                    "price_cents": 1_990,
                    "attributes": {"size": "small"},
                    "is_default": True,
                }
            ],
        ),
    )
    assert created_product.status_code == 201, created_product.text

    exported = admin_client.get("/api/v1/admin/products/export.xlsx")
    assert exported.status_code == 200, exported.text
    workbook = load_workbook(BytesIO(exported.content), read_only=True)
    try:
        assert workbook["Products"].max_row == 2
        assert workbook["SKUs"].max_row == 2
    finally:
        workbook.close()
